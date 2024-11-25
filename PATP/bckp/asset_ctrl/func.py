# funcionalidades de cada tela que for chamada

from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QFrame,QWidget, QLabel, QGraphicsDropShadowEffect, QAbstractItemView
from PyQt5.QtWidgets import QWidget,QPushButton,QFrame,QLineEdit, QComboBox, QDateEdit, QFocusFrame, QScrollArea, QVBoxLayout, QSpinBox, QTableView, QSizePolicy, QHeaderView,QDialog, QListView, QListWidget, QGraphicsView, QGraphicsScene, QFileDialog, QMessageBox, QTextBrowser, QCheckBox
from PyQt5 import uic, QtWidgets, QtCore, QtGui
from PyQt5.QtCore import QResource , QTimer, QLocale, QSortFilterProxyModel, pyqtSignal, QDate
from PyQt5.QtGui import QIcon, QFocusEvent,QDoubleValidator, QStandardItemModel, QStandardItem, QFont
import mysql.connector # type: ignore
from PyQt5.QtGui import QDoubleValidator, QKeyEvent
from PyQt5.QtCore import Qt
import matplotlib.pyplot as plt # type: ignore
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas # type: ignore
from matplotlib.figure import Figure # type: ignore
import pandas as pd
import datetime
from datetime import datetime,timedelta, date
import os, json
import openpyxl #type: ignore
from openpyxl.styles import Font, Alignment #type: ignore
from PyQt5.QtSql import QSqlQueryModel
import traceback
from PyQt5.QtWidgets import QApplication, QTableView, QMainWindow, QVBoxLayout, QWidget, QHeaderView
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from connect import criar_conexao, fechar_conexao, update_editar_ptr, infos_popular_combobox, deletar_ptr,config_acess, config

id_user = ''
data_user = ''
data_pass = ''
data_cargo = 15
if os.path.exists('line/dados.json'):
    print("Arquivo JSON existe.")
    v_j = json.load(open("line/dados.json"))
    id_user = v_j["idusuario"]
    data_user = v_j["user"]
    data_pass = v_j["password"]
    data_cargo = v_j["cargo"]
    print('Usuário:', data_user, 'Senha:', data_pass, 'Cargo:', data_cargo ,'func')
else:
    print("Arquivo JSON inexistente func.")
    
# icones svg
QResource.registerResource("feather/resource.qrc")
home_svg = QIcon("feather/home.svg")
chevrons_left_svg = QIcon("feather/chevrons-left.svg")
menu_svg = QIcon("feather/menu.svg")
user_svg = QIcon("feather/user.svg")
codesandbox_svg = QIcon("feather/codesandbox.svg")
database_svg = QIcon("feather/database.svg")
map_pin_svg = QIcon("feather/map-pin.svg")
dollar_sign_svg = QIcon("feather/dollar-sign.svg")
server_svg = QIcon("feather/server.svg")
search_svg = QIcon("feather/search.svg")
alert_circle_svg = QIcon("feather/alert-circle.svg")
file_text_svg = QIcon("feather/file-text.svg")
users_svg = QIcon("feather/users.svg")
shopping_bag = QIcon("feather/shopping-bag.svg")
activity_svg = QIcon("feather/activity.svg")
save_svg = QIcon("feather/save.svg")
help_circle_svg = QIcon("feather/help-circle.svg")
link_svg = QIcon("feather/link.svg")
user_plus_svg = QIcon("feather/user-plus.svg")
edit_svg = QIcon("feather/edit.svg")
info_svg = QIcon("feather/info.svg")
trash_svg = QIcon("feather/trash.svg")



# Classe do menu usuários, com as funções de cada botão
class user_menu(QWidget):
    def __init__(self):
        super().__init__()
        self.user_main = uic.loadUi("templates/interfaces/user_main.ui", self)
       
        
        self.btn_cad = self.findChild(QPushButton, "btnCad")
        self.btn_cad.setIcon(user_plus_svg)
        self.cad_frame = self.findChild(QFrame, "cadFrame")
        self.btn_cad.installEventFilter(self)

        self.btn_info = self.findChild(QPushButton, "btnDetails")
        self.btn_info.setIcon(info_svg)
        self.info_frame = self.findChild(QFrame, "infoFrame")
        self.btn_info.installEventFilter(self)
        self.btn_info.setVisible(False)        

        self.table_user = self.findChild(QTableView, "l_user")
        self.table_logs = self.findChild(QTableView, "logsReg")

        self.frame_user = self.findChild(QFrame, "info_user")

        self.btn_cad.clicked.connect(self.register)
        self.btn_info.clicked.connect(self.info)
        
        self.shadow_user1 = QGraphicsDropShadowEffect() 
        self.shadow_user1.setOffset(0, 0)
        self.shadow_user1.setBlurRadius(9)
        self.shadow_user1.setColor(QtGui.QColor(0, 0, 0, 128))

        # Reutilizando as configurações de sombra (redução de redundância)
        for i in range(2, 7):
            shadow = QGraphicsDropShadowEffect()
            shadow.setOffset(0, 0)
            shadow.setBlurRadius(9)
            shadow.setColor(QtGui.QColor(0, 0, 0, 128))
            setattr(self, f"shadow_user{i}", shadow)

        self.head_frame = self.findChild(QFrame, "header_user_frame")

        con_u = mysql.connector.connect(**config_acess)
        cursor_u = con_u.cursor()

        # Configuração da tabela de usuários
        cursor_u.execute("""
            SELECT u.usuario, c.nome AS cargo_nome, p.nome AS pessoa_nome, p.dt_create
            FROM usuarios u
            JOIN cargos c ON u.idcargo = c.idcargo
            JOIN pessoas p ON u.idpessoa = p.idpessoa;
        """)
        self.results_u = cursor_u.fetchall()
        model_u = QStandardItemModel(len(self.results_u), 4)
        model_u.setHorizontalHeaderLabels(["Usuário", "Cargo", "Nome", "Data de Criação"])

        for row_idx, row_data in enumerate(self.results_u):
            for col_idx, data in enumerate(row_data):
                item = QStandardItem(str(data))
                item.setFont(QFont("Roboto", 9))
                item.setTextAlignment(Qt.AlignCenter)
                model_u.setItem(row_idx, col_idx, item)

        header = self.table_user.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        self.table_user.setModel(model_u)
        self.table_user.setGridStyle(Qt.NoPen)
        self.table_user.verticalHeader().setVisible(False)
        self.table_user.horizontalHeader().setVisible(True)  # Cabeçalho fixo
        self.table_user.resizeColumnsToContents()
        self.table_user.setAlternatingRowColors(True)
        self.table_user.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        self.table_user.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_user.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_user.clicked.connect(self.handle_row_click)
        select_table = self.table_user.selectionModel()
        select_table.selectionChanged.connect(self.handle_selection_change)

        # Configuração da tabela logsReg
        cursor_u.execute("""
            SELECT c.nome AS cargo_nome, COUNT(u.idpessoa) AS quantidade_usuarios
            FROM usuarios u
            JOIN cargos c ON u.idcargo = c.idcargo
            GROUP BY c.nome
            ORDER BY quantidade_usuarios DESC;
        """)
        self.results_logs = cursor_u.fetchall()
        model_logs = QStandardItemModel(len(self.results_logs), 2)
        model_logs.setHorizontalHeaderLabels(["Cargo", "Quantidade de Usuários"])

        for row_idx, row_data in enumerate(self.results_logs):
            for col_idx, data in enumerate(row_data):
                item = QStandardItem(str(data))
                item.setFont(QFont("Roboto", 9))
                item.setTextAlignment(Qt.AlignCenter)
                model_logs.setItem(row_idx, col_idx, item)

        header_logs = self.table_logs.horizontalHeader()
        header_logs.setSectionResizeMode(QHeaderView.Stretch)
        header_logs.setStretchLastSection(True)
        self.table_logs.setModel(model_logs)
        self.table_logs.setGridStyle(Qt.NoPen)
        self.table_logs.verticalHeader().setVisible(False)
        self.table_logs.horizontalHeader().setVisible(True)  # Cabeçalho fixo
        self.table_logs.resizeColumnsToContents()
        self.table_logs.setAlternatingRowColors(True)
        self.table_logs.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        self.table_logs.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_logs.setEditTriggers(QAbstractItemView.NoEditTriggers)

    # Restante das funções (handle_row_click, handle_selection_change, register, eventFilter, info)
    # Mantidas conforme o código original
    # confirmação do index clickado
    def handle_row_click(self, index):
        self.row = index.row()
        modelo = self.table_user.model()
        item_id = modelo.item(self.row, 0).text()
        print(f"Dados da linha {self.row}: {item_id}")
        self.btn_info.setEnabled(True)
        self.btn_info.show()
        self.btn_info.setVisible(True)
    # seleção
    def handle_selection_change(self, selected, deselected):
        if self.table_user.selectionModel().hasSelection():
            self.btn_info.setEnabled(True)
            self.btn_info.setVisible(True)
        else:
            self.btn_info.setEnabled(False)
            self.btn_info.setVisible(False)
            
    def register(self):
        self.n_frame = self.findChild(QFrame, "new_User")
        self.u_frame = self.findChild(QFrame, "userFrame")
        self.user_create = new_user()
        self.n_frame.layout().addWidget(self.user_create)
        self.u_frame.hide()
        self.n_frame.show()
    
    
    def eventFilter(self, obj, event):

        if obj == self.btn_cad:
            if event.type() == QtCore.QEvent.Enter:
                self.cad_frame.setStyleSheet("QPushButton#btnCad{border:2px solid #666666;border-radius:15px;}")
            elif event.type() == QtCore.QEvent.Leave:
                self.cad_frame.setStyleSheet("QPushButton#btnCad{border: 0px solid transparent;}")

        elif obj == self.btn_info:
            if event.type() == QtCore.QEvent.Enter:
                self.info_frame.setStyleSheet("QPushButton#btnDetails{border:2px solid #666666;border-radius:15px;}")
            elif event.type() == QtCore.QEvent.Leave:
                self.info_frame.setStyleSheet("QPushButton#btnDetails{border: 0px solid transparent;}")

        return super().eventFilter(obj, event)
    
    def info(self):
        modelo = self.table_user.model()
        u_select = modelo.item(self.row, 0).text()
        self.user_edit = edit_user(u_select)
        self.c_frame = self.findChild(QFrame, "createFrame")
        self.u_frame = self.findChild(QFrame, "userFrame")
        self.c_frame.layout().addWidget(self.user_edit)
        self.u_frame.hide()
        self.c_frame.show()
        

class user_info(QWidget):
    def __init__(self, user_id):
        super().__init__()
        self.user_info = uic.loadUi("templates/interfaces/user_info.ui", self)
        self.btn_foto = self.findChild(QPushButton, "btn_foto")
        self.btn_edit = self.findChild(QPushButton, "btn_edit")
        self.btn_view = self.findChild(QPushButton, "view_pass")
        self.cancel_btn = self.findChild(QPushButton, "cancel_btn")
        self.confirm_btn = self.findChild(QPushButton, "confirm_btn")
        self.user_line = self.findChild(QLineEdit, "user_line")
        self.pass_line_info = self.findChild(QLineEdit, "pass_line_info")
        self.name_user_info = self.findChild(QLineEdit, "name_user_info")
        self.email_user_info = self.findChild(QLineEdit, "email_user_info")
        self.cargo_box = self.findChild(QComboBox, "cargo_box")
        self.btn_foto.hide()
        self.btn_view.hide()
        self.cancel_btn.hide()
        self.confirm_btn.hide()
        self.set_fields_enabled(False)
        self.cargo_box.setEnabled(False)
        self.pass_line_info.setEchoMode(QLineEdit.Password)
        self.user_id = user_id
        self.original_data = {}
        self.load_user_data()
        self.btn_edit.clicked.connect(self.enable_edit_mode)
        self.cancel_btn.clicked.connect(self.cancel_changes)
        self.confirm_btn.clicked.connect(self.confirm_changes)
        self.btn_view.clicked.connect(self.toggle_password_view)

    def set_fields_enabled(self, enabled):
        self.user_line.setEnabled(enabled)
        self.pass_line_info.setEnabled(enabled)
        self.name_user_info.setEnabled(enabled)
        self.email_user_info.setEnabled(enabled)

    def load_user_data(self):
        con = criar_conexao()
        cursor = con.cursor()

        try:
            cursor.execute("""
                SELECT u.usuario, u.senha, p.nome, p.email, c.nome AS cargo
                FROM usuarios u
                JOIN pessoas p ON u.idpessoa = p.idpessoa
                JOIN cargos c ON u.idcargo = c.idcargo
                WHERE u.idpessoa = %s
            """, (self.user_id,))
            data = cursor.fetchone()

            if data:
                self.user_line.setText(data[0])
                self.pass_line_info.setText(data[1])
                self.name_user_info.setText(data[2])
                self.email_user_info.setText(data[3])
                self.cargo_box.addItem(data[4])
                self.original_data = {
                    "usuario": data[0],
                    "senha": data[1],
                    "nome": data[2],
                    "email": data[3],
                    "cargo": data[4],
                }
        except Exception as e:
            print(f"Erro ao carregar dados do usuário: {e}")
        finally:
            con.close()

    def enable_edit_mode(self):
        self.set_fields_enabled(True)
        self.btn_foto.show()
        self.btn_view.show()
        self.cancel_btn.show()
        self.confirm_btn.show()
        self.btn_edit.hide()

    def cancel_changes(self):
        self.user_line.setText(self.original_data["usuario"])
        self.pass_line_info.setText(self.original_data["senha"])
        self.name_user_info.setText(self.original_data["nome"])
        self.email_user_info.setText(self.original_data["email"])
        self.set_fields_enabled(False)
        self.btn_foto.hide()
        self.btn_view.hide()
        self.cancel_btn.hide()
        self.confirm_btn.hide()
        self.btn_edit.show()

    def confirm_changes(self):
        usuario = self.user_line.text()
        senha = self.pass_line_info.text()
        nome = self.name_user_info.text()
        email = self.email_user_info.text()

        con = criar_conexao()
        cursor = con.cursor()

        try:
            print('usuario que alterou =', id_user)
            cursor.execute('set @idusuario = %s', (1,))
            cursor.execute("""
                UPDATE usuarios 
                SET usuario = %s, senha = %s
                WHERE idpessoa = %s
            """, (usuario, senha, self.user_id))

            cursor.execute("""
                UPDATE pessoas 
                SET nome = %s, email = %s 
                WHERE idpessoa = %s
            """, (nome, email, self.user_id))

            con.commit()
            self.original_data.update({
                "usuario": usuario,
                "senha": senha,
                "nome": nome,
                "email": email,
            })
            self.cancel_changes()
        except Exception as e:
            print(f"Erro ao atualizar dados do usuário: {e}")
        finally:
            con.close()

    def toggle_password_view(self):

        if self.pass_line_info.echoMode() == QLineEdit.Password:
            self.pass_line_info.setEchoMode(QLineEdit.Normal)
        else:
            self.pass_line_info.setEchoMode(QLineEdit.Password)


class bag_view(QWidget):
    def __init__(self, interface):
        super().__init__()
        self.interface = interface
        self.bag_screen = uic.loadUi("templates/interfaces/bag.ui", self)
        self.produtos = []
        self.produtos_temporarios = []


        self.btn_add_item = self.findChild(QPushButton, "regItem")
        self.btn_add_item.clicked.connect(self.bag_cad)
        self.btn_add_item.installEventFilter(self)


        self.btn_details = self.findChild(QPushButton, "details_btn")
        self.btn_details.installEventFilter(self)
        self.btn_details.hide()



        self.edit_btn = self.findChild(QPushButton, "edit_btn")
        self.edit_btn.clicked.connect(self.bag_edit)

        self.frame_view = self.findChild(QFrame, "frame_view")
        self.frame_search = self.findChild(QFrame, "frame_search")
        self.frame_edit = self.findChild(QFrame, "frame_edit")
        self.frame_details = self.findChild(QFrame, "frame_detail")


        self.shadow_view = QGraphicsDropShadowEffect() 
        self.shadow_view.setOffset(0, 0)
        self.shadow_view.setBlurRadius(9)
        self.shadow_view.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_search = QGraphicsDropShadowEffect() 
        self.shadow_search.setOffset(0, 0)
        self.shadow_search.setBlurRadius(9)
        self.shadow_search.setColor(QtGui.QColor(0, 0, 0, 128))


        self.shadow_edit = QGraphicsDropShadowEffect() 
        self.shadow_edit.setOffset(0, 0)
        self.shadow_edit.setBlurRadius(9)
        self.shadow_edit.setColor(QtGui.QColor(0, 0, 0, 128))


        self.shadow_detail = QGraphicsDropShadowEffect() 
        self.shadow_detail.setOffset(0, 0)
        self.shadow_detail.setBlurRadius(9)
        self.shadow_detail.setColor(QtGui.QColor(0, 0, 0, 128))

        self.frame_view.setGraphicsEffect(self.shadow_view)
        self.frame_search.setGraphicsEffect(self.shadow_search)
        self.frame_edit.setGraphicsEffect(self.shadow_edit)
        self.frame_details.setGraphicsEffect(self.shadow_detail)
        self.search_line = self.findChild(QLineEdit, "line_search")
        self.table_item = self.findChild(QTableView, "table_bag")
        con = mysql.connector.connect(**config)
        cursor = con.cursor()
        cursor.execute("select idpatrimonio,p.nome as patrimonio_nome,p.data_recebimento,c.nome as categoria_nome,s.nome as setor_nome from patrimonios p left join categorias c ON p.idcategoria = c.idcategoria left join setores_responsaveis s ON p.idsetor = s.idsetor order by p.idpatrimonio desc;")
        self.results_mdl = cursor.fetchall()
        self.modelo = QStandardItemModel(len(self.results_mdl), 5)
        for row_idx, row_data in enumerate(self.results_mdl):
            for col_idx, data in enumerate(row_data):
                item = QStandardItem(str(data))
                item.setFont(QFont("Roboto", 9))
                item.setTextAlignment(Qt.AlignCenter)
                self.modelo.setItem(row_idx, col_idx, item)
        header_item = self.table_item.horizontalHeader()
        header_item.setSectionResizeMode(QHeaderView.Stretch)
        header_item.setStretchLastSection(True)
        self.table_item.setModel(self.modelo)        
        self.table_item.verticalHeader().setVisible(False)
        self.table_item.horizontalHeader().setVisible(False)
        self.table_item.resizeColumnsToContents()
        self.table_item.setColumnWidth(1, 250)
        self.table_item.hideColumn(0)
        
        header = self.table_item.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        
        self.table_item.setAlternatingRowColors(True)
        self.table_item.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        
        self.table_item.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_item.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.selected_rows = []
        self.table_item.clicked.connect(self.handle_row_click)
        

        self.l_t = None
        if self.l_t != None:
            self.btn_details.show()
        else:
            self.btn_details.hide()

        self.verif_cargo()
        self.btn_details.clicked.connect(self.exibir_detalhes)
        self.search_line.textChanged.connect(self.filtrar_dados)
        
        self.edit_btn.setEnabled(False)
        self.edit_btn.setVisible(False)
        select_table = self.table_item.selectionModel()
        select_table.selectionChanged.connect(self.handle_selection_change)
        
    def handle_selection_change(self, selected, deselected):
        if self.table_item.selectionModel().hasSelection():
            self.edit_btn.setEnabled(True)
            self.edit_btn.setVisible(True)
        else:
            self.edit_btn.setEnabled(False)
            self.edit_btn.setVisible(False)

    def handle_row_click(self, index):
        self.row = index.row()
        modelo = self.table_item.model()
        item_id = modelo.item(self.row, 0).text()
        print(f"Dados da linha {self.row}: {item_id}")
        self.btn_details.setEnabled(True)
        self.btn_details.show()
        
    def exibir_detalhes(self):
        print('Exibir detalhes: ', self.row)
        modelo_atual = self.table_item.model()
        item_id = modelo_atual.item(self.row, 0).text()
        print(f"ID do item selecionado: {item_id}")

        """Exibir detalhes do item selecionado."""
        if hasattr(self, 'selected_rows'):
            # Recuperar o modelo atualmente exibido
            modelo_atual = self.table_item.model()
            item_id = modelo_atual.item(self.row, 0).text()
            # Recuperar o ID da primeira coluna (mesmo que esteja oculta)
            dialog = detail_window(item_id)
            dialog.exec_()
            print('janela chamada')

    def carregar_dados(self, data, modelo):
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_data in enumerate(row_data):
                item = QStandardItem(str(cell_data))
                item.setFont(QFont("Roboto", 9))
                item.setTextAlignment(Qt.AlignCenter)
                modelo.setItem(row_idx, col_idx, item)
        self.table_item.horizontalHeader().setVisible(False)

    def filtrar_dados(self, texto):
        texto = texto.lower()
        modelo_filtrado = QStandardItemModel(0, 5)
        self.table_item.horizontalHeader().setVisible(False)

        for row in self.results_mdl:
            if any(texto in str(cell).lower() for cell in row):
                items = [QStandardItem(str(cell)) for cell in row]
                for item in items:
                    item.setFont(QFont("Roboto", 9))
                    item.setTextAlignment(Qt.AlignCenter)
                modelo_filtrado.appendRow(items)

        self.table_item.setModel(modelo_filtrado)

    def verif_cargo(self):
        if data_cargo == 4 or data_cargo == 3:
            self.btn_add_item.hide()
            self.edit_btn.hide()        
        else:
            pass
        
    def bag_edit(self):
        try:
            modelo = self.table_item.model()
            item_id = modelo.item(self.row, 0).text()
            self.edit_itens = bag_edit(item_id, self.interface)
            self.edit_frame = self.findChild(QFrame, "frame_edit_2")
            self.body_frame = self.findChild(QFrame, "frame_2")
            self.edit_frame.layout().addWidget(self.edit_itens)
            self.btn_teste = self.findChild(QPushButton, "test")
            self.body_frame.hide()
            self.edit_frame.show()
        except Exception as e:
            print(e)
            QMessageBox.warning(
            self,
            "Aviso",
            "Selecione uma linha antes de continuar!"
        )

    def bag_cad(self):
        self.cad_itens = bag_item_cad()
        self.cad_frame = self.findChild(QFrame, "frame_edit_2")
        self.body_frame = self.findChild(QFrame, "frame_2")
        self.cad_frame.layout().addWidget(self.cad_itens)
        self.btn_teste = self.findChild(QPushButton, "test")
        self.body_frame.hide()
        self.cad_frame.show()     

class bag_edit(QWidget): 
    def __init__(self, id_item, interface):
        super().__init__()
        self.interface = interface
        self.edit = uic.loadUi("templates/interfaces/item_edit.ui", self)
        self.id_item = id_item
        self.n_item = self.findChild(QLabel, "label")
        self.n_item.setText(f"Item: {self.id_item}")
        self.n_f = self.findChild(QLineEdit, "chave_acesso")
        self.n_s = self.findChild(QLineEdit, "n_serie")
        self.n_p = self.findChild(QLineEdit, "n_patr")
        self.fornecedor = self.findChild(QLineEdit, "forn")
        self.name_p = self.findChild(QLineEdit, "name_prod")
        self.v_u = self.findChild(QLineEdit, "value_prod")  
        self.c_sit = self.findChild(QComboBox, "c_sit")
        self.c_set = self.findChild(QComboBox, "c_setor")
        self.c_item = self.findChild(QComboBox, "cat_item")
        self.local_i = self.findChild(QComboBox, 'local_item')
        self.edit_btn = self.findChild(QPushButton, "edit_item")
        self.return_btn = self.findChild(QPushButton, "return_btn")
        self.dt_buy = self.findChild(QDateEdit, "date_buy")
        self.dt_rec = self.findChild(QDateEdit, "date_rec")
        self.del_btn = self.findChild(QPushButton, "btn_del")
        self.num_ptr = self.findChild(QLineEdit, "edtNumPtr")
        self.confirm_item = self.findChild(QPushButton, "confirm_item")
        self.cancel_btn = self.findChild(QPushButton, "cancel_btn")
        
            
        #adicionando os dados do patrimonio selecionado quando abre a tela
        resultado = infos_popular_combobox(self, self.id_item)
                
        self.n_f.setText(resultado[0])
        self.fornecedor.setText(resultado[1])
        self.n_s.setText(resultado[2])
        self.n_p.setText(str(resultado[3]))
        self.dt_buy.setDate(resultado[4])
        self.dt_rec.setDate(resultado[5])
        self.local_i.setCurrentText(resultado[6])
        self.name_p.setText(resultado[7])
        self.v_u.setText(str(resultado[8]))
        self.c_sit.setCurrentText(resultado[9])
        self.c_set.setCurrentText(resultado[10])
        self.c_item.setCurrentText(resultado[11])
        
        #atualizando os valores no banco de dados (update) ao clicar em confirmar
        self.confirm_item.clicked.connect(lambda: 
            update_editar_ptr(
                self.name_p.text(), 
                self.v_u.text(), 
                self.n_s.text(), 
                self.n_p.text(), 
                self.dt_rec.date().toString("yyyy-MM-dd"), 
                self.local_i.currentText(), 
                self.c_sit.currentText(), 
                self.c_set.currentText(), 
                self.c_item.currentText(), 
                self.id_item))
           
        self.confirm_item.clicked.connect(lambda: self.interface.bag_screen())       
        
        #deletando o patrimonio selecionado ao clicar em deletar
        self.del_btn.clicked.connect(lambda: deletar_ptr(self.id_item))

        self.del_btn.clicked.connect(lambda: self.interface.bag_screen())

        #voltar à tela inicial ao clicar em voltar ou cancelar
        self.cancel_btn.clicked.connect(lambda: self.interface.bag_screen())
        self.return_btn.clicked.connect(lambda: self.interface.bag_screen())
        
    def bag_view(self):
        self.interface.bag_screen()
        self.tela_anterior = bag_view()
        self.tela_anterior.configRequested.connect(self.interface.config_screen)
        self.edit_frame = self.findChild(QFrame, "frame")
        self.body_frame = self.findChild(QFrame, "frame_2")
        self.edit_frame.layout().addWidget(self.tela_anterior)
        self.edit_frame.hide()
        self.body_frame.show()
    
class bag_item_cad(QWidget):
    def __init__(self):
        super().__init__()
        
        self.cad_item = uic.loadUi("templates/interfaces/bag_item_cad.ui", self)
        self.name_item = self.findChild(QLineEdit, "name_prod");        
        self.quantidade = self.findChild(QSpinBox, "qnt")
        self.categ_item = self.findChild(QComboBox, "cat_item")
        self.dt_rec = self.findChild(QDateEdit, "date_rec")
        self.setor_item = self.findChild(QComboBox, "c_setor")
        self.sit_item = self.findChild(QComboBox, "c_sit")
        self.rec_por = self.findChild(QLineEdit, "rec_p")
        self.chave_acesso = self.findChild(QLineEdit, "chave_acesso")
        self.n_nota = self.findChild(QLineEdit, "n_nota")
        self.n_serie = self.findChild(QLineEdit, "n_serie")
        self.forn_name = self.findChild(QComboBox, "forn_name")
        self.date_buy = self.findChild(QDateEdit, "date_buy")
        self.date_rec = self.findChild(QDateEdit, "date_rec")
        today = datetime.now()
        self.date_buy.setDate(today)
        self.date_rec.setDate(today)

        self.btn_cancel = self.findChild(QPushButton, "cancelBtn")
        self.btn_cancel.clicked.connect(self.cancel_event)
        #colocando as categorias na combo box
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute('select nome from categorias order by nome')
        resultado = cursor.fetchall()
        for dados in resultado:
            self.categ_item.addItem(dados[0])
            
        cursor.execute('select nome from fornecedores order by nome')
        resultado_forn = cursor.fetchall()
        for dados in resultado_forn:
            self.forn_name.addItem(dados[0])
            
        cursor.execute('select nome from setores_responsaveis order by nome')
        resultado_set_resp = cursor.fetchall()
        for dados in resultado_set_resp:
            self.setor_item.addItem(dados[0])
            
        cursor.execute('select nome from situacoes order by nome')
        resultado_situacoes = cursor.fetchall()
        for dados in resultado_situacoes:
            self.sit_item.addItem(dados[0])
            

        cursor.close()
        fechar_conexao(con)

        # filtro para o campo de valor, aplicada regras para não aceitar virgula e nem outros digitos.
        self.number_input = self.findChild(QLineEdit, "value_prod")
        validator = QDoubleValidator(0.0, 10000.0, 2)
        validator.setNotation(QDoubleValidator.StandardNotation)
        validator.setLocale(self.number_input.locale())
        locale = QLocale(QLocale.C)
        validator.setLocale(locale)
        self.number_input.setValidator(validator)
        self.btn_ok = self.findChild(QPushButton, "ok_btn")
        self.number_input.keyPressEvent = self.keyPressEvent
        self.btn_confirm = self.findChild(QPushButton, "cadItens")
        self.btn_confirm.clicked.connect(self.confirm)

        self.del_btn = self.findChild(QPushButton, "del")
        self.del_btn.clicked.connect(self.deletar_item)

        self.listagem = {}
        self.id_counter = 0
        self.lista_produtos = self.findChild(QTableView, "list_itens_add")

        self.btn_ok.clicked.connect(self.temp_list)

        self.selected_row = None

        self.lista_produtos.clicked.connect(self.handle_row_click)
        self.selected_rows = []
        self.lista_produtos.setGridStyle(Qt.NoPen)
        self.lista_produtos.setAlternatingRowColors(True)
        self.lista_produtos.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        self.lista_produtos.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.lista_produtos.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.value_result = self.findChild(QLabel, "value_r")

        self.number_input.textChanged.connect(self.update_label)
        self.quantidade.valueChanged.connect(self.update_label)

        self.atualizar_tabela()


    def cancel_event(self):
        self.rec_por.clear()
        self.quantidade.clear()
        self.name_item.clear()
        self.number_input.clear()
        self.categ_item.setCurrentIndex(0)
        self.chave_acesso.clear()
        self.n_nota.clear()
        self.n_serie.clear()
        self.forn_name.setCurrentIndex(0)
        self.date_buy.setDate(datetime.now())
        self.date_rec.setDate(datetime.now())
        self.setor_item.setCurrentIndex(0)
        self.sit_item.setCurrentIndex(0)

    def update_label(self):
        try:
            valor = float(self.number_input.text() if self.number_input.text() else 0)
            q = float(self.quantidade.value())
            v = float(valor)
            resultado = q * v
            self.value_result.setText(f"Valor Total: {resultado:.2f}")
        except ValueError:
            self.value_result.setText("Por favor, insira um número válido no campo de valor.")
        pass
    

    def temp_list(self):
        nome = self.name_item.text().strip()
        valor = self.number_input.text()
        categoria = self.categ_item.currentText()
        quantidade = self.quantidade.value()

        data_aquisicao = self.date_buy.date().toString("yyyy-MM-dd")
        chave_acesso = self.chave_acesso.text()
        numero = self.n_nota.text()
        serie = self.n_serie.text()

        print(chave_acesso)

        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute('SELECT idfornecedor FROM fornecedores WHERE nome = %s', (self.forn_name.currentText(),))
        fornecedor_id = cursor.fetchone()[0]
        
        cursor.execute('SELECT nome FROM setores_responsaveis WHERE nome = %s', (self.setor_item.currentText(),))
        setor_nome = cursor.fetchone()[0]
        
        cursor.execute('SELECT nome FROM situacoes WHERE nome = %s', (self.sit_item.currentText(),))
        situacao = cursor.fetchone()[0]
        
        cursor.close()
        fechar_conexao(con)

        if not nome or not valor or quantidade == 0 or not chave_acesso or not numero or not serie or not data_aquisicao:
            print('Faltando valores. Verifique!')
        else:
            self.id_counter += 1
            self.listagem[self.id_counter] = [nome, valor, categoria, quantidade, setor_nome, situacao]
            self.atualizar_tabela()
            self.clear_c()

            # Armazena os dados da nota fiscal
            self.chave_acesso_temp = chave_acesso
            self.numero_nota_temp = numero
            self.serie_nota_temp = serie
            self.data_aquisicao_temp = data_aquisicao
            self.forn_sel_id = fornecedor_id
            # Não insere os itens agora, só guarda na lista local
            print(f'Itens para cadastro: {self.listagem}')

        
    def clear_c(self):
        self.name_item.clear()
        self.number_input.clear()
        self.quantidade.clear()
        self.quantidade.setValue(0)

    def clear_u(self):
        self.name_item.clear()
        self.number_input.clear()
        self.quantidade.clear()
        self.quantidade.setValue(0)
        self.forn_name.setCurrentIndex(0)
        self.setor_item.setCurrentIndex(0)
        self.sit_item.setCurrentIndex(0)
        self.lista_produtos.clearSelection()
        self.listagem.clear()
        self.chave_acesso.clear()
        self.n_nota.clear()
        self.n_serie.clear()
        self.date_buy.setDate(datetime.now())
        self.date_rec.setDate(datetime.now())
        self.rec_por.clear()
        self.atualizar_tabela()

    def atualizar_tabela(self):
        mdl = QStandardItemModel()
        hdr = ["Nome", "Valor", "Categoria", "Quantidade", "Setor", "Situação"]
        mdl.setHorizontalHeaderLabels(hdr)
        for item_id, item_data in self.listagem.items():
            row_items = []
            for value in item_data:
                item = QStandardItem(str(value))
                row_items.append(item)
            mdl.appendRow(row_items)
        self.lista_produtos.setModel(mdl)
        self.lista_produtos.verticalHeader().setVisible(False)
        self.lista_produtos.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.lista_produtos.horizontalHeader().setStretchLastSection(True)
        print(self.listagem)
        
                
    # função responsável pela interação de seleção de linha
    # print com o for apenas para teste
    def handle_row_click(self, index):
        row = index.row()
        self.selected_row = row
        print(f"Dados da linha {row}:")
        for column in range(self.lista_produtos.model().columnCount()):
            data = self.lista_produtos.model().index(row, column).data()
            print(f"{self.lista_produtos.model()}: {data}")
        self.selected_rows.clear()
        self.selected_rows.append(self.lista_produtos.model().index(row, column).data())
        
    def deletar_item(self):
        if self.selected_row is None:
            print("Nenhuma linha selecionada para deletar.")
            return
        item_id = list(self.listagem.keys())[self.selected_row]
        del self.listagem[item_id]
        self.atualizar_tabela()
        self.selected_row = None
        if self.listagem:
            self.id_counter = max(self.listagem.keys())
        else:
            self.id_counter = 0

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == Qt.Key_Comma:
            return
        QLineEdit.keyPressEvent(self.number_input, event)
        


    def confirm(self):
        print("Confirmando os itens:", self.listagem)
        print (self.chave_acesso_temp)
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute('SELECT chave_acesso FROM info_notas WHERE chave_acesso = %s', (self.chave_acesso_temp,))
        chave_acesso_existente = cursor.fetchone()
        cursor.close()
        fechar_conexao(con)

        if chave_acesso_existente:
            print("Nota fiscal ja cadastrada")
            QMessageBox.warning(self, "Erro", "Nota fiscal ja cadastrada")
            self.clear_u()
            return        


        con = criar_conexao()
        cursor = con.cursor()

        try:
            # Insere a nota fiscal na tabela info_notas
            print("Inserindo nota fiscal...")
            cursor.execute('''
                INSERT INTO info_notas (chave_acesso, numero, serie, idfornecedor, data_aquisicao)
                VALUES (%s, %s, %s, %s, %s)
            ''', (self.chave_acesso_temp, self.numero_nota_temp, self.serie_nota_temp, self.forn_sel_id, self.data_aquisicao_temp))
            
            con.commit()
            print("Nota fiscal inserida com sucesso.")

            # Recupera o idnota da nota fiscal recém inserida
            print("Recuperando ID da nota fiscal...")
            cursor.execute('SELECT idnota FROM info_notas WHERE chave_acesso = %s', (self.chave_acesso_temp,))
            id_nota = cursor.fetchone()[0]
            print(f"ID da nota fiscal: {id_nota}")

            # toda a seção for foi usado prints para poder verificar se estava indo certo, e onde poderia ocorrer algum erro ou bug.

            # Processa cada item na listagem
            for item_id, item_data in self.listagem.items():
                try:
                    nome, valor_unitario, categoria, quantidade,setor_nome, situacao = item_data
                    print(f"Processando item: ID={item_id}, Nome={nome}, Valor={valor_unitario}, Categoria={categoria}, Quantidade={quantidade} Setor={setor_nome}, Situação={situacao}")

                    # ID categoria
                    print(f"Buscando ID da categoria para '{categoria}'...")
                    cursor.execute('SELECT idcategoria FROM categorias WHERE nome = %s', (categoria,))
                    resultado_cat = cursor.fetchone()
                    if resultado_cat is None:
                        raise ValueError(f"Categoria '{categoria}' não encontrada no banco de dados.")
                    id_categoria = resultado_cat[0]

                    #id do setor responsável
                    setor_nome = self.setor_item.currentText()
                    print(f"Buscando ID do setor responsável para '{setor_nome}'...")
                    cursor.execute('SELECT idsetor FROM setores_responsaveis WHERE nome = %s', (setor_nome,))
                    resultado_setor = cursor.fetchone()
                    if resultado_setor is None:
                        raise ValueError(f"Setor '{setor_nome}' não encontrado no banco de dados.")
                    id_setor_responsavel = resultado_setor[0]

                    #obter o ID da situação
                    situacao_nome = self.sit_item.currentText()
                    print(f"Buscando ID da situação para '{situacao_nome}'...")
                    cursor.execute('SELECT idsituacao FROM situacoes WHERE nome = %s', (situacao_nome,))
                    resultado_situacao = cursor.fetchone()
                    if resultado_situacao is None:
                        raise ValueError(f"Situação '{situacao_nome}' não encontrada no banco de dados.")
                    id_situacao = resultado_situacao[0]

                    serie_patrimonio = ''
                    # Insere cada unidade do produto individualmente pela quantidade resgatada do item
                    for _ in range(quantidade):
                        cursor.execute('set @idusuario = %s', (id_user,))
                        cursor.execute('''
                            INSERT INTO patrimonios (nome, valor_unitario, data_recebimento, idnota, idcategoria, idsetor, idsituacao, num_serie)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ''', (nome, valor_unitario, self.data_aquisicao_temp, id_nota, id_categoria, id_setor_responsavel, id_situacao, serie_patrimonio))
                    print(f"Item {item_id} inserido com sucesso.")
                # prints do erro
                except Exception as item_error:
                    print(f"Erro ao processar item {item_id}: {item_error}")
                    traceback.print_exc()

            con.commit()
            print(f'{len(self.listagem)} itens e a nota foram cadastrados com sucesso!')

        except Exception as e:
            print('Erro ao cadastrar a nota ou os itens:', e)
            traceback.print_exc()
            con.rollback()

        finally:
            cursor.close()
            fechar_conexao(con)
            self.clear_u()


class items_view(QWidget):
    def __init__(self, interface):
        super().__init__()
        self.interface = interface
        self.item_view = uic.loadUi("templates/interfaces/item_view.ui", self)
        self.tb_item = self.findChild(QTableView, "tb_item")
        self.ct_view = self.findChild(QTableView, "ct_view")
        self.load_items()

    def load_items(self):
        # Conectar ao banco de dados
        conn = mysql.connector.connect(**config)  # Substituir pelo seu dicionário de configuração
        cursor = conn.cursor()

        # Consulta para a tabela tb_item
        tb_item_query = """
        SELECT 
            ptr.nome AS Produto,
            cat.nome AS Categoria,
            sit.nome AS Situação
        FROM 
            patrimonios AS ptr
        INNER JOIN 
            categorias AS cat ON ptr.idcategoria = cat.idcategoria
        INNER JOIN 
            situacoes AS sit ON ptr.idsituacao = sit.idsituacao;
        """

        # Consulta para a tabela ct_view (quantidade de produtos por categoria)
        ct_view_query = """
        SELECT 
            cat.nome AS Categoria,
            COUNT(ptr.idpatrimonio) AS Quantidade
        FROM 
            categorias AS cat
        LEFT JOIN 
            patrimonios AS ptr ON cat.idcategoria = ptr.idcategoria
        GROUP BY 
            cat.nome
        ORDER BY 
            cat.nome;
        """

        # Executar e preencher tb_item
        cursor.execute(tb_item_query)
        tb_item_results = cursor.fetchall()
        self.setup_table(self.tb_item, tb_item_results, ['Produto', 'Categoria', 'Situação'])

        # Executar e preencher ct_view
        cursor.execute(ct_view_query)
        ct_view_results = cursor.fetchall()
        self.setup_table(self.ct_view, ct_view_results, ['Categoria', 'Quantidade'])

        # Fechar o cursor e a conexão
        cursor.close()
        conn.close()

    def setup_table(self, table_view, data, headers):
        """Configura o QTableView com os dados fornecidos."""
        # Criar o modelo
        model = QStandardItemModel(len(data), len(headers))
        model.setHorizontalHeaderLabels(headers)

        # Preencher o modelo com os dados
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_data in enumerate(row_data):
                item = QStandardItem(str(cell_data))
                item.setFont(QFont("Roboto", 9))
                item.setTextAlignment(Qt.AlignCenter)
                model.setItem(row_idx, col_idx, item)

        # Configurar o QTableView
        table_view.setModel(model)
        table_view.resizeColumnsToContents()
        table_view.setAlternatingRowColors(True)
        table_view.setStyleSheet(
            "alternate-background-color: #F0F0F0; background-color: #FFFFFF;"
        )
        table_view.setEditTriggers(QAbstractItemView.NoEditTriggers)  # Desabilitar edição
        table_view.setSelectionMode(QAbstractItemView.NoSelection)  # Desabilitar seleção
        table_view.horizontalHeader().setStretchLastSection(True)
        table_view.verticalHeader().setVisible(False)  # Ocultar cabeçalho vertical


    def setup_table(self, table_view, data, headers):
        """Configura o QTableView com os dados fornecidos."""
        # Criar o modelo
        model = QStandardItemModel(len(data), len(headers))
        model.setHorizontalHeaderLabels(headers)

        # Preencher o modelo com os dados
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_data in enumerate(row_data):
                item = QStandardItem(str(cell_data))
                item.setFont(QFont("Roboto", 9))
                item.setTextAlignment(Qt.AlignCenter)
                model.setItem(row_idx, col_idx, item)

        # Configurar o QTableView
        table_view.setModel(model)
        table_view.resizeColumnsToContents()
        table_view.setAlternatingRowColors(True)
        table_view.setStyleSheet(
            "alternate-background-color: #F0F0F0; background-color: #FFFFFF;"
        )
        table_view.setEditTriggers(QAbstractItemView.NoEditTriggers)  # Desabilitar edição
        table_view.setSelectionMode(QAbstractItemView.NoSelection)  # Desabilitar seleção
        table_view.horizontalHeader().setStretchLastSection(True)
        table_view.verticalHeader().setVisible(False)  # Ocultar cabeçalho vertical




        
class categ_view(QWidget):   
    configRequested = pyqtSignal()
    def __init__(self):
        super().__init__()
        self.item_category = uic.loadUi("templates/interfaces/categ_view.ui", self)
        self.btn_config = self.findChild(QPushButton, "category")
        self.btn_config.clicked.connect(self.on_config_click)
    def on_config_click(self):
        self.configRequested.emit()

class rel_view(QWidget):
    def __init__(self):
        super().__init__()
        self.rel_view = uic.loadUi("templates/interfaces/rel_view.ui", self)
        self.date_i = self.findChild(QDateEdit, "date_i")
        self.date_f = self.findChild(QDateEdit, "date_f")
        self.set_default_dates()

    def set_default_dates(self):
        today = datetime.now()
        three_years_ago = today - timedelta(days=365 * 1)
        self.date_i.setDate(three_years_ago)
        self.date_f.setDate(today)


class patr_view(QWidget):
    def __init__(self):
        super().__init__()
        self.patr_view = uic.loadUi("templates/interfaces/patr_view.ui", self)
        self.list_ptr = self.findChild(QTableView, "list_patr")
        self.vl_ctg = self.findChild(QTableView, "vl_ctg")
        self.vl_setor = self.findChild(QTableView, "vl_setor")
        self.grap_view = self.findChild(QGraphicsView, "grap_view")
        self.date_i = self.findChild(QDateEdit, "dt_inicial")
        self.date_f = self.findChild(QDateEdit, "dt_final")
        self.btn_filter = self.findChild(QPushButton, "filter_dt")
        self.btn_rlt = self.findChild(QPushButton, "rlt_btn")
        self.lbl_total = self.findChild(QLabel, "vl_total")
        self.lbl_patr = self.findChild(QLabel, "vl_ptr")
        self.frame_header = self.findChild(QFrame, "frame_2")

        self.shadow_frame1 = QGraphicsDropShadowEffect() 
        self.shadow_frame1.setOffset(0, 0)  # posição da sombra
        self.shadow_frame1.setBlurRadius(9)  # tamanho da sombra
        self.shadow_frame1.setColor(QtGui.QColor(0, 100, 0, 128))

        self.shadow_frame2 = QGraphicsDropShadowEffect() 
        self.shadow_frame2.setOffset(0, 0)  # posição da sombra
        self.shadow_frame2.setBlurRadius(9)  # tamanho da sombra
        self.shadow_frame2.setColor(QtGui.QColor(0, 100, 0, 128))

        self.shadow_frame3 = QGraphicsDropShadowEffect() 
        self.shadow_frame3.setOffset(0, 0)  # posição da sombra
        self.shadow_frame3.setBlurRadius(9)  # tamanho da sombra
        self.shadow_frame3.setColor(QtGui.QColor(0, 100, 0, 128))

        self.shadow_frame4 = QGraphicsDropShadowEffect() 
        self.shadow_frame4.setOffset(0, 0)  # posição da sombra
        self.shadow_frame4.setBlurRadius(9)  # tamanho da sombra
        self.shadow_frame4.setColor(QtGui.QColor(0, 100, 0, 128))
        
        self.list_ptr.setGraphicsEffect(self.shadow_frame1)
        self.vl_ctg.setGraphicsEffect(self.shadow_frame2)
        self.vl_setor.setGraphicsEffect(self.shadow_frame3)
        self.frame_header.setGraphicsEffect(self.shadow_frame4)

        self.set_default_dates()
        self.btn_filter.clicked.connect(self.filter_data)
        self.btn_rlt.clicked.connect(self.rlt_patrimonio)
        self.filter_data()



    def set_default_dates(self):
        today = datetime.now()
        three_years_ago = today - timedelta(days=365 * 3)
        self.date_i.setDate(three_years_ago)
        self.date_f.setDate(today)

    def filter_data(self):
        """Filtra e carrega os dados para as tabelas."""
        d_i = self.date_i.date().toString("yyyy-MM-dd")
        d_f = self.date_f.date().toString("yyyy-MM-dd")

        try:
            con = criar_conexao()
            cursor = con.cursor(dictionary=True)

            # Query para a tabela de patrimônio
            cursor.execute("""
                SELECT 
                    p.nome, 
                    p.valor_unitario, 
                    p.data_recebimento, 
                    p.num_patrimonio
                FROM patrimonios p
                WHERE p.data_recebimento BETWEEN %s AND %s
            """, (d_i, d_f))
            patrimonios = cursor.fetchall()

            if patrimonios:
                self.load_table_data(self.list_ptr, patrimonios)
            else:
                self.list_ptr.setModel(None)
                QMessageBox.information(self, "Aviso", "Nenhum item encontrado no intervalo de datas.")

            # Query para valores por categoria
            cursor.execute("""
                SELECT 
                    c.nome AS categoria, 
                    SUM(p.valor_unitario) AS total_valor
                FROM categorias c
                JOIN patrimonios p ON p.idcategoria = c.idcategoria
                WHERE p.data_recebimento BETWEEN %s AND %s
                GROUP BY c.nome
            """, (d_i, d_f))
            categorias = cursor.fetchall()

            self.load_table_data(self.vl_ctg, categorias)

            # Query para valores por setor
            cursor.execute("""
                SELECT 
                    sr.nome AS setor, 
                    SUM(p.valor_unitario) AS total_valor
                FROM setores_responsaveis sr
                JOIN patrimonios p ON p.idsetor = sr.idsetor
                WHERE p.data_recebimento BETWEEN %s AND %s
                GROUP BY sr.nome
            """, (d_i, d_f))
            setores = cursor.fetchall()

            self.load_table_data(self.vl_setor, setores)

            # Calcular valores totais
            cursor.execute("""
                SELECT 
                    SUM(valor_unitario) AS total, 
                    SUM(CASE WHEN num_patrimonio IS NOT NULL THEN valor_unitario ELSE 0 END) AS total_patrimonio
                FROM patrimonios
                WHERE data_recebimento BETWEEN %s AND %s
            """, (d_i, d_f))
            totais = cursor.fetchone()

            self.lbl_total.setText(f"Valor total dos produtos registrados: R$ {totais['total']:.2f}")
            self.lbl_patr.setText(f"Valor total dos patrimônios: R$ {totais['total_patrimonio']:.2f}")

        except Exception as e:
            QMessageBox.critical(self, "Valores vazios", f"Ocorreu um erro ao realizar a busca, nenhum item existe no intervalo de datas: {e}")
        finally:
            con.close()


    def load_table_data(self, table, data):
        """Carrega os dados em uma tabela QTableView e ajusta automaticamente, centralizando o texto."""
        model = QStandardItemModel()
        if not data:
            # Caso não tenha dados, insira uma mensagem
            model.setHorizontalHeaderLabels(["Aviso"])
            model.appendRow([QStandardItem("Nenhum item encontrado.")])
            table.setModel(model)
            return

        # Definir cabeçalhos com base nas chaves do dicionário
        headers = data[0].keys()
        model.setHorizontalHeaderLabels(headers)

        # Preencher os dados
        for row in data:
            items = []
            for value in row.values():
                item = QStandardItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)  # Centralizar o texto
                items.append(item)
            model.appendRow(items)

        # Aplicar o modelo na tabela
        table.setModel(model)

        # Ajustar tamanho das colunas de forma responsiva ao tamanho da tabela
        header = table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)

        # Remover cabeçalhos horizontal e vertical
        table.horizontalHeader().setVisible(False)
        table.verticalHeader().setVisible(False)

        # Aplicar estilos
        table.setAlternatingRowColors(True)  # Habilitar cores alternadas
        table.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")

        # Ajustar tamanho das linhas
        table.resizeRowsToContents()


    def rlt_patrimonio(self):
        """Gera o relatório em Excel."""
        d_i = self.date_i.date().toString("yyyy-MM-dd")
        d_f = self.date_f.date().toString("yyyy-MM-dd")

        file_path, _ = QFileDialog.getSaveFileName(self, "Salvar Relatório", os.path.expanduser("~"), "Arquivos Excel (*.xlsx)")
        if not file_path:
            return

        try:
            con = criar_conexao()
            cursor = con.cursor(dictionary=True)

            workbook = openpyxl.Workbook()

            # Adicionar dados de patrimônio
            cursor.execute("""
                SELECT 
                    p.nome AS nome_patrimonio, 
                    p.valor_unitario, 
                    p.num_patrimonio, 
                    p.data_recebimento
                FROM patrimonios p
                WHERE p.data_recebimento BETWEEN %s AND %s
            """, (d_i, d_f))
            patrimonios = cursor.fetchall()

            patr_sheet = workbook.active
            patr_sheet.title = "Patrimônios"
            self.add_sheet_data(patr_sheet, patrimonios)

            # Adicionar dados de categorias
            cursor.execute("""
                SELECT 
                    c.nome AS categoria, 
                    SUM(p.valor_unitario) AS total_valor
                FROM categorias c
                JOIN patrimonios p ON p.idcategoria = c.idcategoria
                WHERE p.data_recebimento BETWEEN %s AND %s
                GROUP BY c.nome
            """, (d_i, d_f))
            categorias = cursor.fetchall()

            cat_sheet = workbook.create_sheet("Categorias")
            self.add_sheet_data(cat_sheet, categorias)

            # Adicionar dados de setores
            cursor.execute("""
                SELECT 
                    sr.nome AS setor, 
                    SUM(p.valor_unitario) AS total_valor
                FROM setores_responsaveis sr
                JOIN patrimonios p ON p.idsetor = sr.idsetor
                WHERE p.data_recebimento BETWEEN %s AND %s
                GROUP BY sr.nome
            """, (d_i, d_f))
            setores = cursor.fetchall()

            setor_sheet = workbook.create_sheet("Setores")
            self.add_sheet_data(setor_sheet, setores)

            # Salvar o arquivo Excel
            workbook.save(file_path)
            QMessageBox.information(self, "Relatório Gerado", f"O relatório foi salvo em: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro: {e}")
        finally:
            con.close()


    def add_sheet_data(self, sheet, data):
        """Adiciona dados a uma planilha do Excel."""
        if data:
            headers = list(data[0].keys())
            sheet.append(headers)

            for row in data:
                sheet.append(list(row.values()))
        else:
            sheet.append(["Nenhum dado encontrado."])

class logs_view(QWidget):
    def __init__(self):
        super().__init__()
        self.logs_view = uic.loadUi("templates/interfaces/logs_view.ui", self)

class config_view(QWidget):
    def __init__(self):
        super().__init__()
        self.config_view = uic.loadUi("templates/interfaces/config_screen.ui", self)
    




class local_info(QWidget):
    def __init__(self):
        super().__init__()
        self.local_info = uic.loadUi("templates/interfaces/local_info.ui", self)
        self.local_items = self.findChild(QTableView, 'local_items')
        self.local_noitems = self.findChild(QTableView, 'itens_nlocal')

        # Carregar os dados ao inicializar a tela
        self.load_local_data()

    def load_local_data(self):
        """Carrega os dados das tabelas `local_items` e `itens_nlocal`."""
        # Conectar ao banco de dados
        try:
            conn = mysql.connector.connect(**config)  # Configuração do banco
            cursor = conn.cursor()

            # Consulta para itens com locais e setores vinculados
            local_items_query = """
            SELECT 
                p.nome AS Item,
                l.nome AS Local,
                sr.nome AS Setor
            FROM 
                patrimonios p
            INNER JOIN 
                locais l ON p.idlocal = l.idlocal
            INNER JOIN 
                setores_responsaveis sr ON p.idsetor = sr.idsetor;
            """

            # Consulta para itens sem locais vinculados
            no_local_items_query = """SELECT p.nome AS Item, c.nome AS Categoria FROM patrimonios p INNER JOIN categorias c ON p.idcategoria = c.idcategoria WHERE p.idlocal IS NULL;"""
            # Executar consultas
            cursor.execute(local_items_query)
            local_items_results = cursor.fetchall()

            cursor.execute(no_local_items_query)
            no_local_items_results = cursor.fetchall()

            # Configurar tabelas com os dados
            self.setup_table(
                self.local_items,
                local_items_results,
                ['Item', 'Local', 'Setor'],
                empty_message="Sem dados para apresentar na tabela 'Itens com Local e Setor'."
            )
            self.setup_table(
                self.local_noitems,
                no_local_items_results,
                ['Item', 'Categoria'],
                empty_message="Sem dados para apresentar na tabela 'Itens sem Local'."
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao carregar os dados: {e}")
        finally:
            # Fechar conexão com o banco de dados
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

    def setup_table(self, table_view, data, headers, empty_message):
        """Configura uma QTableView com dados ou exibe uma mensagem em caso de ausência."""
        if not table_view:  # Verifica se o objeto da tabela é válido
            QMessageBox.critical(self, "Erro", "Tabela não encontrada na interface.")
            return

        # Cria o modelo com o número de linhas/colunas
        model = QStandardItemModel(len(data) if data else 1, len(headers))
        model.setHorizontalHeaderLabels(headers)

        if not data:  # Sem dados
            for col_idx, header in enumerate(headers):
                item = QStandardItem(empty_message if col_idx == 0 else "")
                item.setFont(QFont("Roboto", 9, QFont.StyleItalic))
                item.setTextAlignment(Qt.AlignCenter)
                model.setItem(0, col_idx, item)
        else:  # Preenche os dados
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_data in enumerate(row_data):
                    item = QStandardItem(str(cell_data))
                    item.setFont(QFont("Roboto", 9))
                    item.setTextAlignment(Qt.AlignCenter)
                    model.setItem(row_idx, col_idx, item)

        # Configurar o QTableView
        table_view.setModel(model)
        table_view.setAlternatingRowColors(True)
        table_view.setStyleSheet(
            "alternate-background-color: #F0F0F0; background-color: #FFFFFF;"
        )
        table_view.setEditTriggers(QAbstractItemView.NoEditTriggers)  # Desabilitar edição
        table_view.setSelectionMode(QAbstractItemView.NoSelection)  # Desabilitar seleção
        table_view.verticalHeader().setVisible(False)  # Ocultar cabeçalho vertical

        # Configurar o cabeçalho horizontal
        header = table_view.horizontalHeader()
        header.setStretchLastSection(True)  # Última coluna expansível
        header.setSectionsMovable(False)  # Fixar cabeçalho
        header.setSectionsClickable(False)  # Desativar clique no cabeçalho
        header.setSectionResizeMode(QHeaderView.Fixed)  # Tamanhos fixos
        table_view.resizeColumnsToContents()  # Ajustar colunas ao conteúdo inicial

        
        
class detail_window(QDialog):
    def __init__(self, item_id):
        super().__init__()
        print('item_id teste instancia', item_id)
        self.item_id = item_id

        self.setWindowFlags(Qt.Popup| Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_DeleteOnClose)  # Garante que o diálogo será descartado ao fechar
        self.setFixedSize(400, 300)
        self.setStyleSheet("background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 10px;")

        layout = QVBoxLayout()

        try:
            # Conexão com banco de dados
            con = mysql.connector.connect(**config)
            cursor = con.cursor()
            cursor.execute("SELECT * FROM patrimonios WHERE idpatrimonio = %s", (self.item_id,))
            data = cursor.fetchall()
        except mysql.connector.Error as e:
            print("Erro ao conectar ao banco de dados:", e)
            data = []

        if not data:
            print("Nenhum dado foi encontrado na tabela.")
            detail_text = "Nenhum dado foi encontrado na tabela."
        else:
            detail_text = ""
            for row in data:
                detail_text += f"ID: {row[0]}\n"
                detail_text += f"Nome: {row[1]}\n"
                detail_text += f"Valor Unitário: {row[2]}\n"
                detail_text += f"Data de Recebimento: {row[3]}\n"
                detail_text += f"Patrimônio: {row[4]}\n"
        label_mensagem = QLabel(detail_text)
        label_mensagem.setWordWrap(True)
        label_mensagem.setAlignment(Qt.AlignTop)
        layout.addWidget(label_mensagem)

        self.setLayout(layout)

    def mousePressEvent(self, event):
        """Fecha o diálogo se clicar fora dele."""
        if not self.rect().contains(event.pos()):
            self.close()




class config_cargo(QWidget):
    def __init__(self):
        super().__init__()
        self.config_cargo = uic.loadUi("templates/interfaces/cargo_config.ui", self)
        self.box_cargo = self.findChild(QComboBox, "box_cargo")
        self.new_cargo = self.findChild(QLineEdit, "new_cargo")
        self.btn_confirm = self.findChild(QPushButton, "btn_confirm_edit")
        self.btn_cancel = self.findChild(QPushButton, "btn_cancel")
        self.edit_carg = self.findChild(QPushButton, "edit_cargo")
        self.cancel_edit = self.findChild(QPushButton, "btn_edit_cancel")
        self.perm_acess = self.findChild(QCheckBox, "acesso_geral")
        self.perm_reg = self.findChild(QCheckBox, "pode_registrar")
        self.ctrl_adm = self.findChild(QCheckBox, "controle_adm")
        self.ctrl_user = self.findChild(QCheckBox, "controle_usuario")
        self.perm_mod = self.findChild(QCheckBox, "pode_modificar")
        self.perm_read = self.findChild(QCheckBox, "pode_visualizar")
        self.btn_confirm_n = self.findChild(QPushButton, "btn_confirm")
        self.edit_carg.setVisible(False)
        self.btn_confirm_n.hide()
        self.new_cargo.hide()
        self.btn_confirm.hide()
        self.btn_cancel.hide()
        self.cancel_edit.hide()
        self.btn_add = self.findChild(QPushButton, "add_cargo")
        self.box_cargo.setPlaceholderText("Selecione um cargo")
        # Conexão com o banco para adicionar valores a combobox
        self.con = criar_conexao()
        self.cursor = self.con.cursor()
        self.cursor.execute("SELECT nome FROM cargos")
        data = self.cursor.fetchall()
        for cargo in data:
            self.box_cargo.addItem(cargo[0])
        # Inicialmente, checkboxes desabilitadas
        self.set_checkboxes_enabled(False)

        # Conecta sinais
        self.box_cargo.currentIndexChanged.connect(self.update_checkboxes)
        self.edit_carg.clicked.connect(self.enable_editing)
        self.btn_confirm.clicked.connect(self.save_changes)
        self.cancel_edit.clicked.connect(self.cancel_changes)

        # Armazenar os valores originais para restaurar em caso de cancelamento
        self.original_values = {}
        self.btn_add.clicked.connect(self.test)
        self.btn_add.clicked.connect(self.start_new_cargo)
        self.btn_confirm_n.clicked.connect(self.save_new_cargo)
        self.btn_cancel.clicked.connect(self.cancel_new_cargo)

    def start_new_cargo(self):
        # sub-sessão para criação de um novo cargo
        self.btn_add.setVisible(False)
        self.box_cargo.setVisible(False)
        self.new_cargo.show()
        self.btn_confirm_n.show()
        self.btn_cancel.show()
        self.edit_carg.hide()
        self.box_cargo.setEnabled(False)
        self.clear_checkboxes()
        self.set_checkboxes_enabled(True)

    def cancel_new_cargo(self):
        """Cancelar o processo de criação de um novo cargo e restaurar o estado inicial."""
        self.new_cargo.hide()
        self.btn_confirm_n.hide()
        self.edit_carg.show()
        self.box_cargo.setEnabled(True)
        self.box_cargo.setVisible(True)
        self.clear_checkboxes()
        self.set_checkboxes_enabled(False)
        self.box_cargo.setCurrentIndex(0)
        self.btn_add.setVisible(True)

    def save_new_cargo(self):
        """Salvar o novo cargo no banco de dados."""
        cargo_name = self.new_cargo.text().strip()
        if not cargo_name:
            print("Nome do cargo não pode ser vazio.")
            return

        # Verifica se o cargo já existe
        self.cursor.execute("SELECT COUNT(*) FROM cargos WHERE nome = %s", (cargo_name,))
        if self.cursor.fetchone()[0] > 0:
            print("Cargo já existe!")
            return

        # Obtem os valores das permissões do banco em uma tupla
        data = (
            int(self.perm_acess.isChecked()),
            int(self.perm_reg.isChecked()),
            int(self.ctrl_adm.isChecked()),
            int(self.ctrl_user.isChecked()),
            int(self.perm_mod.isChecked()),
            int(self.perm_read.isChecked()),
            cargo_name,
        )

        # Inserir o novo cargo no banco
        self.cursor.execute('set @idusuario = %s', (id_user,))
        self.cursor.execute("""
            INSERT INTO cargos 
            (acesso_geral, pode_registrar, controle_adm, controle_usuario, pode_modificar, pode_visualizar, nome) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, data)
        self.con.commit()

        # atualiza a combobox
        self.box_cargo.addItem(cargo_name)

        # Restaura a interface ao estado inicial
        self.cancel_new_cargo()    


    def test(self):
        cargo = self.box_cargo.currentText()
        self.cursor.execute("""SELECT acesso_geral, pode_registrar, controle_adm, controle_usuario, pode_modificar, pode_visualizar FROM cargos WHERE nome = %s""", (cargo,))
        data = self.cursor.fetchone()
        print(data)

    def set_checkboxes_enabled(self, enabled):
        """Habilitar ou desabilitar as checkboxes."""

        self.perm_acess.setEnabled(enabled)
        self.perm_reg.setEnabled(enabled)
        self.ctrl_adm.setEnabled(enabled)
        self.ctrl_user.setEnabled(enabled)
        self.perm_mod.setEnabled(enabled)
        self.perm_read.setEnabled(enabled)

    def update_checkboxes(self):
        self.clear_checkboxes()  # Garante que as checkbox estejam limpas
        cargo = self.box_cargo.currentText()
        self.edit_carg.setVisible(True)
        if not cargo or cargo == "Selecione um cargo":
            self.set_checkboxes_enabled(False)
            return
        # verificação dos dados da combobox no banco para realizar as atribuições dos checkbox
        try:
            self.cursor.execute("""
                SELECT acesso_geral, pode_registrar, controle_adm, controle_usuario, 
                    pode_modificar, pode_visualizar
                FROM cargos 
                WHERE nome = %s
            """, (cargo,))
            data = self.cursor.fetchone()

            if data:
                # debugar os dados brutos recebidos, utilizando print
                print(f"Dados crus do banco para cargo {cargo}: {data}")

                # Converte os valores de CHAR(1) para inteiros
                safe_data = [int(value) if value is not None else 0 for value in data]

                # Atualizar as checkboxes com os valores convertidos
                self.perm_acess.setChecked(bool(safe_data[0]))
                self.perm_reg.setChecked(bool(safe_data[1]))
                self.ctrl_adm.setChecked(bool(safe_data[2]))
                self.ctrl_user.setChecked(bool(safe_data[3]))
                self.perm_mod.setChecked(bool(safe_data[4]))
                self.perm_read.setChecked(bool(safe_data[5]))

                # Debug o estado das checkboxes com print
                print(f"Estados das checkboxes para {cargo}: "
                    f"perm_acess={self.perm_acess.isChecked()}, "
                    f"perm_reg={self.perm_reg.isChecked()}, "
                    f"ctrl_adm={self.ctrl_adm.isChecked()}, "
                    f"ctrl_user={self.ctrl_user.isChecked()}, "
                    f"perm_mod={self.perm_mod.isChecked()}, "
                    f"perm_read={self.perm_read.isChecked()}")
            else:
                print(f"Nenhum dado encontrado para o cargo {cargo}")
                self.set_checkboxes_enabled(False)
        except Exception as e:
            print(f"Erro ao buscar dados do cargo: {e}")
            self.clear_checkboxes()
            self.set_checkboxes_enabled(False)



    def clear_checkboxes(self):
        self.perm_acess.setChecked(False)
        self.perm_reg.setChecked(False)
        self.ctrl_adm.setChecked(False)
        self.ctrl_user.setChecked(False)
        self.perm_mod.setChecked(False)
        self.perm_read.setChecked(False)


    def enable_editing(self):
        # Armazenar os valores originais antes de permitir a edição
        self.btn_add.setVisible(False)
        self.original_values = {
            "acesso_geral": self.perm_acess.isChecked(),
            "pode_registrar": self.perm_reg.isChecked(),
            "controle_adm": self.ctrl_adm.isChecked(),
            "controle_usuario": self.ctrl_user.isChecked(),
            "pode_modificar": self.perm_mod.isChecked(),
            "pode_visualizar": self.perm_read.isChecked(),
        }

        self.set_checkboxes_enabled(True)
        self.btn_confirm.show()
        self.cancel_edit.show()
        self.edit_carg.hide()


    def save_changes(self):
        """Salvar as alterações no banco de dados."""
        cargo = self.box_cargo.currentText()
        data = (
            int(self.perm_acess.isChecked()),
            int(self.perm_reg.isChecked()),
            int(self.ctrl_adm.isChecked()),
            int(self.ctrl_user.isChecked()),
            int(self.perm_mod.isChecked()),
            int(self.perm_read.isChecked()),
            cargo,
        )
        self.cursor.execute('set @idusuario = %s', (id_user,))
        self.cursor.execute("""
            UPDATE cargos 
            SET acesso_geral = %s, pode_registrar = %s, controle_adm = %s, controle_usuario = %s, pode_modificar = %s, pode_visualizar = %s
            WHERE nome = %s
        """, data)
        self.con.commit()
        self.set_checkboxes_enabled(False)
        self.btn_confirm.hide()
        self.cancel_edit.hide()
        self.edit_carg.show()
        self.box_cargo.setCurrentIndex(0)
        self.btn_add.setVisible(True)
    def cancel_changes(self):
        """Restaurar os valores originais em caso de cancelamento."""
        if not self.original_values:
            print("Nenhum valor original armazenado para restaurar.")
            return

        # restaura os valores originais
        self.perm_acess.setChecked(bool(self.original_values.get("acesso_geral", False)))
        self.perm_reg.setChecked(bool(self.original_values.get("pode_registrar", False)))
        self.ctrl_adm.setChecked(bool(self.original_values.get("controle_adm", False)))
        self.ctrl_user.setChecked(bool(self.original_values.get("controle_usuario", False)))
        self.perm_mod.setChecked(bool(self.original_values.get("pode_modificar", False)))
        self.perm_read.setChecked(bool(self.original_values.get("pode_visualizar", False)))

        # retorna para o layout normal da sessão
        self.set_checkboxes_enabled(False)
        self.btn_confirm.hide()
        self.cancel_edit.hide()
        self.edit_carg.show()
        

class config_cat(QWidget):
    def __init__(self):
        super().__init__()
        self.config_cat = uic.loadUi("templates/interfaces/categ_config.ui", self)

        # Widgets
        self.box_cat = self.findChild(QComboBox, "box_cat")
        self.btn_confirm = self.findChild(QPushButton, "btn_confirm")
        self.btn_add = self.findChild(QPushButton, "btn_add")
        self.btn_cancel = self.findChild(QPushButton, "btn_cancel")
        self.btn_edit = self.findChild(QPushButton, "btn_edit")
        self.line_cat = self.findChild(QLineEdit, "line_cat")
        self.lbl_qntd = self.findChild(QLabel, "lbl_qntd")

        # Estado atual (edição ou adição)
        self.current_mode = None  # None, "edit", ou "add"

        # Inicializar estado
        self.initialize_state()

        # Carregar categorias no comboBox
        self.load_categories()

        # Conectar eventos
        self.box_cat.currentIndexChanged.connect(self.handle_category_selection)
        self.btn_edit.clicked.connect(self.enable_edit_mode)
        self.btn_add.clicked.connect(self.add_category_mode)
        self.btn_confirm.clicked.connect(self.confirm_action)
        self.btn_cancel.clicked.connect(self.cancel_action)

    def initialize_state(self):
        """Define o estado inicial da interface."""
        self.line_cat.setEnabled(False)
        self.btn_confirm.setEnabled(False)
        self.btn_edit.setEnabled(False)
        self.btn_add.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.btn_edit.setVisible(False)
        self.btn_cancel.setVisible(False)
        self.btn_confirm.setVisible(False)
        self.line_cat.clear()
        self.lbl_qntd.clear()
        self.current_mode = None

    def load_categories(self):
        """Carrega as categorias do banco para o comboBox."""
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute("SELECT nome FROM categorias")
        categories = cursor.fetchall()
        con.close()

        self.box_cat.clear()
        self.box_cat.addItem("Selecione uma categoria")
        for category in categories:
            self.box_cat.addItem(category[0])

    def handle_category_selection(self):
        """Atualiza a interface com base na categoria selecionada."""
        selected_index = self.box_cat.currentIndex()
        if selected_index == 0:  # Índice 0 é o placeholder
            self.initialize_state()
            return

        selected_category = self.box_cat.currentText()
        con = criar_conexao()
        cursor = con.cursor(dictionary=True)

        try:
            # Obter dados da categoria selecionada
            cursor.execute("SELECT idcategoria, nome FROM categorias WHERE nome = %s", (selected_category,))
            category_data = cursor.fetchone()

            if not category_data:
                QMessageBox.warning(self, "Erro", "Categoria não encontrada no banco de dados!")
                self.initialize_state()
                return

            # Obter a quantidade de itens vinculados
            cursor.execute("SELECT COUNT(*) AS qtd FROM patrimonios WHERE idcategoria = %s", (category_data["idcategoria"],))
            item_count = cursor.fetchone()["qtd"]

            # Atualizar interface
            self.line_cat.setText(category_data["nome"])
            self.lbl_qntd.setText(str(item_count))
            self.line_cat.setEnabled(False)
            self.btn_edit.setEnabled(True)
            self.btn_confirm.setEnabled(False)
            self.btn_cancel.setEnabled(False)
            self.btn_edit.setVisible(True)
            self.btn_cancel.setVisible(False)
            self.btn_confirm.setVisible(False)
            self.btn_add.setEnabled(True)
            self.btn_add.setVisible(True)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro: {e}")
        finally:
            con.close()

    def enable_edit_mode(self):
        """Ativa o modo de edição."""
        self.line_cat.setEnabled(True)
        self.btn_edit.setEnabled(False)
        self.btn_add.setEnabled(False)
        self.btn_confirm.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_confirm.setVisible(True)
        self.btn_cancel.setVisible(True)
        self.btn_edit.setVisible(False)
        self.current_mode = "edit"  # Define o modo como edição

    def add_category_mode(self):
        """Ativa o modo de adição de categoria."""
        self.line_cat.setEnabled(True)
        self.line_cat.clear()
        self.lbl_qntd.clear()
        self.btn_add.setEnabled(False)
        self.btn_edit.setEnabled(False)
        self.btn_confirm.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_confirm.setVisible(True)
        self.btn_cancel.setVisible(True)
        self.btn_edit.setVisible(False)
        self.current_mode = "add"  # Define o modo como adição

    def confirm_action(self):
        """Confirma a ação (edição ou adição)."""
        new_category_name = self.line_cat.text().strip()

        if not new_category_name:
            QMessageBox.warning(self, "Aviso", "O nome da categoria não pode ser vazio!")
            return

        con = criar_conexao()
        cursor = con.cursor()

        try:
            if self.current_mode == "edit":  # Edição
                current_category_name = self.box_cat.currentText()
                cursor.execute('set @idusuario = %s', (id_user,))
                cursor.execute("UPDATE categorias SET nome = %s WHERE nome = %s", (new_category_name, current_category_name))
            elif self.current_mode == "add":  # Adição
                cursor.execute('set @idusuario = %s', (id_user,))
                cursor.execute("INSERT INTO categorias (nome) VALUES (%s)", (new_category_name,))

            con.commit()
            QMessageBox.information(self, "Sucesso", "Ação realizada com sucesso!")
            self.load_categories()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro: {e}")
        finally:
            con.close()

        self.initialize_state()

    def cancel_action(self):
        """Cancela a ação e reseta o estado inicial."""
        self.initialize_state()


class config_forn(QWidget):
    def __init__(self):
        super().__init__()
        self.config_forn = uic.loadUi("templates/interfaces/forn_config.ui", self)
        self.box_forn = self.findChild(QComboBox, "box_forn")
        self.name_line = self.findChild(QLineEdit, "name_line")
        self.cnpj_line = self.findChild(QLineEdit, "cnpj_line")
        self.confirm_btn = self.findChild(QPushButton, "confirm_btn")
        self.cancel_btn = self.findChild(QPushButton, "cancel_btn")
        self.edit_btn = self.findChild(QPushButton, "edit_btn")
        self.del_btn = self.findChild(QPushButton, "del_btn")
        self.add_btn = self.findChild(QPushButton, "add_forn")
        self.frame_forn = self.findChild(QFrame, "frame_4")
        self.txt_lbl = self.findChild(QLabel, "lbl_text")
        self.txt_lbl.hide()
        self.box_forn.setPlaceholderText("Selecione um fornecedor")
        # Inicialização do estado da interface
        self.confirm_btn.hide()
        self.cancel_btn.hide()
        self.del_btn.hide()
        self.edit_btn.hide()
        self.name_line.setEnabled(False)
        self.cnpj_line.setEnabled(False)

        self.load_fornecedores()

        # Conectando os sinais
        self.box_forn.currentIndexChanged.connect(self.atribui_forn)
        self.edit_btn.clicked.connect(self.enable_edit)
        self.cancel_btn.clicked.connect(self.cancel_edit)
        self.confirm_btn.clicked.connect(self.confirm_changes)
        self.del_btn.clicked.connect(self.delete_forn)
        self.add_btn.clicked.connect(self.add_new_forn)

    def load_fornecedores(self):
        # atualiza a lista de fornecedores na combobox
        con = mysql.connector.connect(**config)
        cursor = con.cursor()
        cursor.execute("SELECT nome FROM fornecedores")
        data = cursor.fetchall()
        con.close()
        self.box_forn.clear()
        for forn in data:
            self.box_forn.addItem(forn[0])

    def atribui_forn(self):
        self.edit_btn.show()
        # Busca os dados do fornecedor selecionado
        con = mysql.connector.connect(**config)
        cursor = con.cursor()
        cursor.execute("SELECT nome, cnpj FROM fornecedores WHERE nome = %s", (self.box_forn.currentText(),))
        resultado = cursor.fetchone()
        con.close()
        if resultado:
            self.name_line.setText(resultado[0])
            self.cnpj_line.setText(resultado[1])
        else:
            self.name_line.clear()
            self.cnpj_line.clear()

    def enable_edit(self):
       # edição de fornecedor
        self.edit_btn.setEnabled(False)
        self.edit_btn.setVisible(False)
        self.add_btn.setEnabled(False)
        self.add_btn.setVisible(False)
        self.confirm_btn.show()
        self.cancel_btn.show()
        self.del_btn.show()
        self.name_line.setEnabled(True)
        self.cnpj_line.setEnabled(True)
        self.box_forn.setEnabled(False)

    def cancel_edit(self):
        #cancela a edição e retorna ao estado inicial
        self.edit_btn.setEnabled(True)
        self.edit_btn.setVisible(True)
        self.add_btn.setEnabled(True)
        self.add_btn.setVisible(True)
        self.confirm_btn.hide()
        self.cancel_btn.hide()
        self.del_btn.hide()
        self.name_line.setEnabled(False)
        self.cnpj_line.setEnabled(False)
        self.box_forn.setVisible(True)
        self.atribui_forn()  # Recarrega os dados originais

    def confirm_changes(self):
        con = mysql.connector.connect(**config)
        cursor = con.cursor()
        try:
            cursor.execute('set @idusuario = %s', (id_user,))
            cursor.execute(
                "UPDATE fornecedores SET nome = %s, cnpj = %s WHERE nome = %s",
                (self.name_line.text(), self.cnpj_line.text(), self.box_forn.currentText())
            )
            con.commit()
        finally:
            con.close()
        self.cancel_edit()
        self.load_fornecedores()

    def delete_forn(self):
        con = mysql.connector.connect(**config)
        cursor = con.cursor()
        try:
            cursor.execute(
                """
                SELECT COUNT(*) FROM info_notas 
                WHERE idfornecedor = (SELECT idfornecedor FROM fornecedores WHERE nome = %s)
                """,
                (self.box_forn.currentText(),)
            )
            vinculos = cursor.fetchone()[0]
            if vinculos > 0:
                QMessageBox.warning(self, "Erro", "Não é possível deletar o fornecedor, pois ele possui notas fiscais vinculadas.")
            else:
                cursor.execute(
                    "DELETE FROM fornecedores WHERE nome = %s",
                    (self.box_forn.currentText(),)
                )
                con.commit()
                QMessageBox.information(self, "Sucesso", "Fornecedor deletado com sucesso.")
                self.cancel_edit()
                self.load_fornecedores()
        finally:
            con.close()

    def add_new_forn(self):
        self.box_forn.setVisible(False) 
        self.name_line.clear()
        self.cnpj_line.clear()
        self.name_line.setEnabled(True)
        self.cnpj_line.setEnabled(True)
        self.add_btn.setVisible(False) 
        self.edit_btn.setEnabled(False)
        self.edit_btn.setVisible(False)
        self.confirm_btn.show()
        self.cancel_btn.show()
        self.del_btn.hide()

        def confirm_add():
            con = mysql.connector.connect(**config)
            cursor = con.cursor()
            try:
                cursor.execute('set @idusuario = %s', (id_user,))
                cursor.execute("INSERT INTO fornecedores (nome, cnpj) VALUES (%s, %s)",(self.name_line.text(), self.cnpj_line.text()))
                con.commit()
                QMessageBox.information(self, "Sucesso", "Novo fornecedor adicionado com sucesso.")
            finally:
                con.close()
            self.cancel_edit()  # Retorna ao estado inicial
            self.box_forn.setVisible(True)
            self.add_btn.setVisible(True)
            self.load_fornecedores()

        # Substitui o comportamento padrão do botão Confirmar
        self.confirm_btn.clicked.disconnect()
        self.confirm_btn.clicked.connect(confirm_add)


        

class config_local(QWidget):
    def __init__(self):
        super().__init__()
        self.config_local = uic.loadUi("templates/interfaces/local_config.ui", self)
        self.box_local = self.findChild(QComboBox, "box_local")
        self.btn_edit = self.findChild(QPushButton, "edit_btn")
        self.btn_add = self.findChild(QPushButton, "add_btn")
        self.btn_cancel = self.findChild(QPushButton, "cancel_btn")
        self.btn_del = self.findChild(QPushButton, "del_btn")
        self.btn_confirm = self.findChild(QPushButton, "confirm_btn")
        self.local_line = self.findChild(QLineEdit, "local_line")
        self.lbl_qntd = self.findChild(QLabel, "lbl_qntd")

        # Configuração inicial
        self.local_line.setEnabled(False)
        self.disable_all_buttons()
        self.load_combobox()

        # Conectar sinais
        self.box_local.currentIndexChanged.connect(self.update_local_data)
        self.btn_edit.clicked.connect(self.enable_edit_mode)
        self.btn_add.clicked.connect(self.enable_add_mode)
        self.btn_cancel.clicked.connect(self.cancel_changes)
        self.btn_confirm.clicked.connect(self.confirm_changes)

    def disable_all_buttons(self):
        """Desabilita todos os botões exceto os básicos."""
        self.btn_edit.setEnabled(False)
        self.btn_add.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.btn_del.setEnabled(False)
        self.btn_confirm.setEnabled(False)
        self.btn_edit.setVisible(False)
        self.btn_cancel.setVisible(False)
        self.btn_add.setVisible(True)
        self.btn_confirm.setVisible(False)
        self.btn_del.setVisible(False)
    def load_combobox(self):
        """Carrega os locais na QComboBox."""
        self.box_local.clear()
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute("SELECT nome FROM locais")
        locais = cursor.fetchall()
        con.close()

        self.box_local.addItem("Selecione um local")
        for local in locais:
            self.box_local.addItem(local[0])

    def update_local_data(self):
        """Atualiza os dados ao selecionar um local na combo-box."""
        local_selecionado = self.box_local.currentText()
        if local_selecionado == "Selecione um local":
            self.local_line.clear()
            self.lbl_qntd.clear()
            self.disable_all_buttons()
            return

        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute("SELECT idlocal, nome FROM locais WHERE nome = %s", (local_selecionado,))
        local_data = cursor.fetchone()

        if not local_data:
            # Caso o local não seja encontrado
            QMessageBox.warning(self, "Banco atualizado", "Novo local adicionado a caixa de seleção.")
            self.local_line.clear()
            self.lbl_qntd.clear()
            self.disable_all_buttons()
            con.close()
            return

        # Obter a quantidade de itens vinculados ao local
        cursor.execute("SELECT COUNT(*) FROM patrimonios WHERE idlocal = %s", (local_data[0],))
        quantidade = cursor.fetchone()[0]
        con.close()

        # Preencher campos
        self.local_line.setText(local_data[1])
        self.lbl_qntd.setText(str(quantidade))

        # Atualizar botões
        self.local_line.setEnabled(False)
        self.btn_edit.setEnabled(True)
        self.btn_del.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.btn_confirm.setEnabled(False)
        self.btn_edit.setVisible(True)


    def enable_edit_mode(self):
        """Habilita o modo de edição."""
        self.local_line.setEnabled(True)
        self.btn_edit.setEnabled(False)
        self.btn_add.setEnabled(False)
        self.btn_confirm.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_confirm.setVisible(True)
        self.btn_cancel.setVisible(True)
        self.btn_edit.setVisible(False)
        

    def enable_add_mode(self):
        """Habilita o modo de adição de um novo local."""
        self.local_line.clear()
        self.local_line.setEnabled(True)
        self.btn_edit.setEnabled(False)
        self.btn_add.setEnabled(False)
        self.btn_del.setEnabled(False)
        self.btn_confirm.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_confirm.setVisible(True)
        self.btn_cancel.setVisible(True)
        self.btn_edit.setVisible(False)

    def cancel_changes(self):
        """Cancela alterações e retorna ao estado inicial."""
        self.update_local_data()  # Recarregar dados do local selecionado
        self.local_line.setEnabled(False)
        self.btn_edit.setEnabled(True)
        self.btn_add.setEnabled(True)
        self.btn_confirm.setEnabled(False)
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.setVisible(False)
        self.btn_confirm.setVisible(False)
        self.btn_edit.setVisible(True)

    def confirm_changes(self):
        """Confirma alterações ou adição de um novo local."""
        novo_nome = self.local_line.text().strip()
        if not novo_nome:
            QMessageBox.warning(self, "Atenção", "O nome do local não pode estar vazio.")
            return

        local_selecionado = self.box_local.currentText()

        con = criar_conexao()
        cursor = con.cursor()

        try:
            if local_selecionado == "Selecione um local":
                # Adicionar novo local
                cursor.execute('set @idusuario = %s', (id_user,))
                cursor.execute("INSERT INTO locais (nome) VALUES (%s)", (novo_nome,))
            else:
                # Editar local existente
                cursor.execute('set @idusuario = %s', (id_user,))
                cursor.execute("UPDATE locais SET nome = %s WHERE nome = %s", (novo_nome, local_selecionado))

            con.commit()
            QMessageBox.information(self, "Sucesso", "Alteração confirmada!")
            self.load_combobox()
        except Exception as e:
            con.rollback()
            QMessageBox.critical(self, "Erro", f"Falha ao salvar alterações: {e}")
        finally:
            con.close()

        self.cancel_changes()

class edit_user(QWidget):
    def __init__(self, user_edit):
        super().__init__()
        self.user_edit = user_edit  # Agora representa o nome do usuário
        self.user_info = uic.loadUi("templates/interfaces/user_edit.ui", self)
        self.btn_foto = self.findChild(QPushButton, "btn_foto")
        self.btn_edit = self.findChild(QPushButton, "btn_edit")
        self.btn_view = self.findChild(QPushButton, "view_pass")
        self.cancel_btn = self.findChild(QPushButton, "cancel_btn")
        self.confirm_btn = self.findChild(QPushButton, "confirm_btn")
        self.user_line = self.findChild(QLineEdit, "user_line")
        self.pass_line_info = self.findChild(QLineEdit, "pass_line_info")
        self.name_user_info = self.findChild(QLineEdit, "name_user_info")
        self.email_user_info = self.findChild(QLineEdit, "email_user_info")
        self.cargo_box = self.findChild(QComboBox, "cargo_box")
        self.lbl_user = self.findChild(QLabel, "label")
        self.lbl_user.setText(self.user_edit)
        self.btn_foto.hide()
        self.btn_view.hide()
        self.cancel_btn.hide()
        self.confirm_btn.hide()
        self.set_fields_enabled(False)
        self.pass_line_info.setEchoMode(QLineEdit.Password)
        self.original_data = {}
        self.load_user_data()
        self.load_cargos()  # Carregar cargos
        self.btn_edit.clicked.connect(self.enable_edit_mode)
        self.cancel_btn.clicked.connect(self.cancel_changes)
        self.confirm_btn.clicked.connect(self.confirm_changes)
        self.btn_view.clicked.connect(self.toggle_password_view)

    def set_fields_enabled(self, enabled):
        self.user_line.setEnabled(enabled)
        self.pass_line_info.setEnabled(enabled)
        self.name_user_info.setEnabled(enabled)
        self.email_user_info.setEnabled(enabled)
        self.cargo_box.setEnabled(enabled)  # Habilitar a edição do cargo

    def load_user_data(self):
        """Carrega os dados do usuário com base no nome do usuário."""
        con = criar_conexao()
        cursor = con.cursor()

        try:
            # Ajusta a consulta para buscar pelo nome do usuário
            cursor.execute("""
                SELECT u.usuario, u.senha, p.nome, p.email, c.nome AS cargo, u.idcargo
                FROM usuarios u
                JOIN pessoas p ON u.idpessoa = p.idpessoa
                JOIN cargos c ON u.idcargo = c.idcargo
                WHERE u.usuario = %s
            """, (self.user_edit,))
            data = cursor.fetchone()

            if data:
                self.user_line.setText(data[0])
                self.pass_line_info.setText(data[1])
                self.name_user_info.setText(data[2])
                self.email_user_info.setText(data[3])
                self.cargo_box.setCurrentText(data[4])  # Preencher o ComboBox com o cargo
                self.original_data = {
                    "usuario": data[0],
                    "senha": data[1],
                    "nome": data[2],
                    "email": data[3],
                    "cargo": data[4],
                    "idcargo": data[5],  # Armazenar o ID do cargo
                }
            else:
                QMessageBox.warning(self, "Aviso", "Usuário não encontrado no banco de dados.")
        except Exception as e:
            print(f"Erro ao carregar dados do usuário: {e}")
        finally:
            con.close()

    def load_cargos(self):
        """Carrega os cargos disponíveis para o ComboBox."""
        con = criar_conexao()
        cursor = con.cursor()

        try:
            cursor.execute("SELECT nome FROM cargos")
            cargos = cursor.fetchall()

            # Adicionar cargos ao ComboBox
            for cargo in cargos:
                self.cargo_box.addItem(cargo[0])
        except Exception as e:
            print(f"Erro ao carregar cargos: {e}")
        finally:
            con.close()

    def enable_edit_mode(self):
        self.set_fields_enabled(True)
        self.btn_foto.show()
        self.btn_view.show()
        self.cancel_btn.show()
        self.confirm_btn.show()
        self.btn_edit.hide()

    def cancel_changes(self):
        self.user_line.setText(self.original_data["usuario"])
        self.pass_line_info.setText(self.original_data["senha"])
        self.name_user_info.setText(self.original_data["nome"])
        self.email_user_info.setText(self.original_data["email"])
        self.cargo_box.setCurrentText(self.original_data["cargo"])  # Reverter o cargo
        self.set_fields_enabled(False)
        self.btn_foto.hide()
        self.btn_view.hide()
        self.cancel_btn.hide()
        self.confirm_btn.hide()
        self.btn_edit.show()

    def confirm_changes(self):
        usuario = self.user_line.text()
        senha = self.pass_line_info.text()
        nome = self.name_user_info.text()
        email = self.email_user_info.text()
        cargo = self.cargo_box.currentText()  # Obter o cargo selecionado

        con = criar_conexao()
        cursor = con.cursor()

        try:
            # Atualizar os dados do usuário
            cursor.execute("""
                UPDATE usuarios 
                SET usuario = %s, senha = %s, idcargo = (
                    SELECT idcargo FROM cargos WHERE nome = %s
                )
                WHERE usuario = %s
            """, (usuario, senha, cargo, self.user_edit))

            cursor.execute("""
                UPDATE pessoas 
                SET nome = %s, email = %s 
                WHERE idpessoa = (
                    SELECT idpessoa FROM usuarios WHERE usuario = %s
                )
            """, (nome, email, self.user_edit))

            con.commit()
            self.original_data.update({
                "usuario": usuario,
                "senha": senha,
                "nome": nome,
                "email": email,
                "cargo": cargo,
            })
            self.cancel_changes()
        except Exception as e:
            print(f"Erro ao atualizar dados do usuário: {e}")
        finally:
            con.close()

    def toggle_password_view(self):
        if self.pass_line_info.echoMode() == QLineEdit.Password:
            self.pass_line_info.setEchoMode(QLineEdit.Normal)
        else:
            self.pass_line_info.setEchoMode(QLineEdit.Password)


class new_user(QWidget):
    def __init__(self):
        super().__init__()
        self.user_create = uic.loadUi("templates/interfaces/user_create.ui", self)
        self.cargo_box = self.findChild(QComboBox, "cargoBox")
        self.user_edit_line = self.findChild(QLineEdit, "userEdit")
        self.name_edit_line = self.findChild(QLineEdit, "nameEdit")
        self.pass_edit_line = self.findChild(QLineEdit, "passEdit")
        self.email_edit_line = self.findChild(QLineEdit, "emailEdit")
        self.btn_clear = self.findChild(QPushButton, "btnClear")
        self.btn_confirm = self.findChild(QPushButton, "btnConfirm")

        self.cargo_att()
        self.btn_clear.clicked.connect(self.clear)
        self.btn_confirm.clicked.connect(self.create_user)

    def cargo_att(self):
        """Popula o ComboBox com os cargos disponíveis no banco."""
        con_cargo = mysql.connector.connect(**config)
        cursor2 = con_cargo.cursor()
        try:
            query = "SELECT nome FROM cargos"
            cursor2.execute(query)
            dados = cursor2.fetchall()
            for dado in dados:
                self.cargo_box.addItem(dado[0])
        except Exception as e:
            print(f"Erro ao carregar cargos: {e}")
        finally:
            con_cargo.close()

    def create_user(self):
        """Cria um novo usuário nas tabelas pessoas e usuarios."""
        user = self.user_edit_line.text()
        name = self.name_edit_line.text()
        passw = self.pass_edit_line.text()
        email = self.email_edit_line.text()
        cargo = self.cargo_box.currentText()

        if not user or not name or not passw or not email:
            print("Preencha todos os campos.")
            return

        con = mysql.connector.connect(**config)
        cursor = con.cursor()

        try:
            # Verifica se o nome de usuário já existe
            cursor.execute("SELECT COUNT(*) FROM usuarios WHERE usuario = %s", (user,))
            if cursor.fetchone()[0] > 0:
                print("Nome de usuário já existe.")
                return

            # Obtém o ID do cargo selecionado
            cursor.execute("SELECT idcargo FROM cargos WHERE nome = %s", (cargo,))
            cargo_id = cursor.fetchone()
            if not cargo_id:
                print("Cargo inválido.")
                return

            cargo_id = cargo_id[0]

            # Insere os dados na tabela pessoas
            dt_create = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute('set @idusuario = %s', (id_user,))
            cursor.execute(
                "INSERT INTO pessoas (nome, email, dt_create) VALUES (%s, %s, %s)",
                (name, email, dt_create),
            )
            pessoa_id = cursor.lastrowid

            # Insere os dados na tabela usuarios
            cursor.execute(
                "INSERT INTO usuarios (idpessoa, usuario, senha, idcargo) VALUES (%s, %s, %s, %s)",
                (pessoa_id, user, passw, cargo_id),
            )

            con.commit()
            print("Usuário criado com sucesso.")
            self.clear()

        except Exception as e:
            print(f"Erro ao criar usuário: {e}")
            con.rollback()
        finally:
            con.close()

    def clear(self):
        """Limpa os campos de entrada."""
        self.user_edit_line.clear()
        self.name_edit_line.clear()
        self.pass_edit_line.clear()
        self.email_edit_line.clear()
        print("Campos limpos.")
