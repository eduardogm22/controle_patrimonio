# funcionalidades de cada tela que for chamada

from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QFrame,QWidget, QLabel, QGraphicsDropShadowEffect, QAbstractItemView
from PyQt5.QtWidgets import QWidget,QPushButton,QFrame,QLineEdit, QComboBox, QDateEdit, QFocusFrame, QScrollArea, QVBoxLayout, QSpinBox, QTableView, QSizePolicy, QHeaderView,QDialog, QListView, QListWidget, QGraphicsView, QGraphicsScene, QFileDialog, QMessageBox
from PyQt5 import uic, QtWidgets, QtCore, QtGui
from PyQt5.QtCore import QResource , QTimer, QLocale, QSortFilterProxyModel, pyqtSignal
from PyQt5.QtGui import QIcon, QFocusEvent,QDoubleValidator, QStandardItemModel, QStandardItem
from connect import conecta_view_tela, conecta_procedure_tela, criar_conexao, fechar_conexao, config_acess, config, cadastra_nota, deletar_patrimonio
import mysql.connector # type: ignore
from PyQt5.QtGui import QDoubleValidator, QKeyEvent
from PyQt5.QtCore import Qt
import matplotlib.pyplot as plt # type: ignore
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas # type: ignore
from matplotlib.figure import Figure # type: ignore
import pandas as pd
from datetime import datetime,timedelta
import os, json
import openpyxl #type: ignore
from openpyxl.styles import Font, Alignment #type: ignore
import time

time.sleep(1)
id_user = ''
data_user = ''
data_pass = ''
data_cargo = 15
if os.path.exists('line/dados.json'):
    print("Arquivo JSON existe.")
    v_j = json.load(open("line/dados.json"))
    id_user = v_j["idusuario"]
    data_user = v_j["user"]
    data_pass = v_j["password"]
    data_cargo = v_j["cargo"]
    print('ID Usuário', id_user,'Usuário:', data_user, 'Senha:', data_pass, 'Cargo:', data_cargo ,'func')
else:
    print("Arquivo JSON inexistente func.")

# icones svg
QResource.registerResource("feather/resource.qrc")
home_svg = QIcon("feather/home.svg")
chevrons_left_svg = QIcon("feather/chevrons-left.svg")
menu_svg = QIcon("feather/menu.svg")
user_svg = QIcon("feather/user.svg")
codesandbox_svg = QIcon("feather/codesandbox.svg")
database_svg = QIcon("feather/database.svg")
map_pin_svg = QIcon("feather/map-pin.svg")
dollar_sign_svg = QIcon("feather/dollar-sign.svg")
server_svg = QIcon("feather/server.svg")
search_svg = QIcon("feather/search.svg")
alert_circle_svg = QIcon("feather/alert-circle.svg")
file_text_svg = QIcon("feather/file-text.svg")
users_svg = QIcon("feather/users.svg")
shopping_bag = QIcon("feather/shopping-bag.svg")
activity_svg = QIcon("feather/activity.svg")
save_svg = QIcon("feather/save.svg")
help_circle_svg = QIcon("feather/help-circle.svg")
link_svg = QIcon("feather/link.svg")
user_plus_svg = QIcon("feather/user-plus.svg")
edit_svg = QIcon("feather/edit.svg")
info_svg = QIcon("feather/info.svg")
trash_svg = QIcon("feather/trash.svg")



# Classe do menu usuários, com as funções de cada botão
class user_menu(QWidget):
    def __init__(self):
        super().__init__()
        self.user_main = uic.loadUi("templates/interfaces/user_main.ui", self)
        self.btn_search = self.findChild(QPushButton,"btnSearch")
        self.btn_search.setIcon(search_svg)
        self.search_frame = self.findChild(QFrame, "search_frame")
        self.btn_search.installEventFilter(self)        
        
        self.btn_cad = self.findChild(QPushButton, "btnCad")
        self.btn_cad.setIcon(user_plus_svg)
        self.cad_frame = self.findChild(QFrame, "cadFrame")
        self.btn_cad.installEventFilter(self)

        self.btn_info = self.findChild(QPushButton, "btnDetails")
        self.btn_info.setIcon(info_svg)
        self.info_frame = self.findChild(QFrame, "infoFrame")
        self.btn_info.installEventFilter(self)
        
        self.table_user = self.findChild(QTableView, "userList")
        self.table_logs = self.findChild(QTableView, "logsReg")
        self.search_line = self.findChild(QLineEdit, "search_input")
        
        self.frame_user = self.findChild(QFrame, "info_user")

        self.btn_search.clicked.connect(self.search)
        self.btn_cad.clicked.connect(self.register)
        self.btn_info.clicked.connect(self.info)
        
        self.shadow_user1 = QGraphicsDropShadowEffect() 
        self.shadow_user1.setOffset(0, 0)
        self.shadow_user1.setBlurRadius(9)
        self.shadow_user1.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_user2 = QGraphicsDropShadowEffect() 
        self.shadow_user2.setOffset(0, 0)
        self.shadow_user2.setBlurRadius(9)
        self.shadow_user2.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_user3 = QGraphicsDropShadowEffect() 
        self.shadow_user3.setOffset(0, 0)
        self.shadow_user3.setBlurRadius(9)
        self.shadow_user3.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_user4 = QGraphicsDropShadowEffect() 
        self.shadow_user4.setOffset(0, 0)
        self.shadow_user4.setBlurRadius(9)
        self.shadow_user4.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_user5 = QGraphicsDropShadowEffect() 
        self.shadow_user5.setOffset(0, 0)
        self.shadow_user5.setBlurRadius(9)
        self.shadow_user5.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_user6 = QGraphicsDropShadowEffect() 
        self.shadow_user6.setOffset(0, 0)
        self.shadow_user6.setBlurRadius(9)
        self.shadow_user6.setColor(QtGui.QColor(0, 0, 0, 128))

        self.head_frame = self.findChild(QFrame, "header_user_frame")

        self.search_frame.setGraphicsEffect(self.shadow_user1)
        self.info_frame.setGraphicsEffect(self.shadow_user5)
        self.cad_frame.setGraphicsEffect(self.shadow_user2)
        self.frame_user.setGraphicsEffect(self.shadow_user3)
        self.head_frame.setGraphicsEffect(self.shadow_user6)


            
    def register(self):
        self.c_frame = self.findChild(QFrame, "createFrame")
        self.u_frame = self.findChild(QFrame, "userFrame")
        self.user_create = uic.loadUi("templates/interfaces/user_create.ui")
        self.c_frame.layout().addWidget(self.user_create)
        self.btn_clear = self.findChild(QPushButton, "btnClear")
        self.btn_clear.clicked.connect(self.clear)
        self.btn_confirm = self.findChild(QPushButton, "btnConfirm")
        self.btn_confirm.clicked.connect(self.create_user)
        self.frame_bot = self.findChild(QFrame, "frame_bottom")
        self.frame_body = self.findChild(QFrame, "frame_body")
        self.frame_rodape = self.findChild(QFrame, "frame_rodape")
        self.cargo_box = self.findChild(QComboBox, "cargoBox")
        self.cargo_box.setStyleSheet("QComboBox#cargoBox{border: 1px solid #000000;} QComboBox#cargoBox::drop-down{border: 1px solid #000000; border-radius:9px;}")
        
        self.user_edit_line = self.findChild(QLineEdit, "userEdit")
        self.name_edit_line = self.findChild(QLineEdit, "nameEdit")
        self.pass_edit_line = self.findChild(QLineEdit, "passEdit")
        self.email_edit_line = self.findChild(QLineEdit, "emailEdit")

        self.frame_btns = self.findChild(QFrame, "frame_btns")
        self.frame_back = self.findChild(QFrame, "frame_back")

        self.btn_details = self.findChild(QPushButton, "btn_details")
        self.btn_details.setIcon(alert_circle_svg)
        
        self.shadow_details = QGraphicsDropShadowEffect() 
        self.shadow_details.setOffset(0, 0)
        self.shadow_details.setBlurRadius(9)
        self.shadow_details.setColor(QtGui.QColor(0, 0, 0, 128))
    
        self.shadow_f1 = QGraphicsDropShadowEffect() 
        self.shadow_f1.setOffset(0, 0)
        self.shadow_f1.setBlurRadius(9)
        self.shadow_f1.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_f2 = QGraphicsDropShadowEffect() 
        self.shadow_f2.setOffset(0, 0)
        self.shadow_f2.setBlurRadius(9)
        self.shadow_f2.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_f3 = QGraphicsDropShadowEffect() 
        self.shadow_f3.setOffset(0, 0)
        self.shadow_f3.setBlurRadius(9)
        self.shadow_f3.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_f4 = QGraphicsDropShadowEffect() 
        self.shadow_f4.setOffset(0, 0)
        self.shadow_f4.setBlurRadius(9)
        self.shadow_f4.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_f5 = QGraphicsDropShadowEffect() 
        self.shadow_f5.setOffset(0, 0)
        self.shadow_f5.setBlurRadius(9)
        self.shadow_f5.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_f6 = QGraphicsDropShadowEffect() 
        self.shadow_f6.setOffset(0, 0)
        self.shadow_f6.setBlurRadius(9)
        self.shadow_f6.setColor(QtGui.QColor(0, 0, 0, 128))

        self.shadow_f7 = QGraphicsDropShadowEffect() 
        self.shadow_f7.setOffset(0, 0)
        self.shadow_f7.setBlurRadius(9)
        self.shadow_f7.setColor(QtGui.QColor(0, 0, 0, 128))

        self.user_edit_line.setGraphicsEffect(self.shadow_f1)
        self.name_edit_line.setGraphicsEffect(self.shadow_f2)
        self.pass_edit_line.setGraphicsEffect(self.shadow_f3)
        self.email_edit_line.setGraphicsEffect(self.shadow_f4)
        self.cargo_box.setGraphicsEffect(self.shadow_f5)
        self.frame_btns.setGraphicsEffect(self.shadow_f6)
        self.frame_back.setGraphicsEffect(self.shadow_f7)

        self.frame_bot.setGraphicsEffect(self.shadow_details)
        self.cargo_box.setStyleSheet("QComboBox#cargoBox{text-align: center;}")
        self.u_frame.hide()
        self.c_frame.show()
        con_cargo = mysql.connector.connect(**config)
        cursor2 = con_cargo.cursor()
        query = "SELECT nome FROM cargos"
        cursor2.execute(query)
        dados = cursor2.fetchall()
        for dado in dados:
            self.cargo_box.addItem(dado[0])
        self.u_frame.hide()
        self.c_frame.show()
        con_cargo.close()
    
    def create_user(self):
        user = self.user_edit_line.text()
        name = self.name_edit_line.text()
        passw = self.pass_edit_line.text()
        email = self.email_edit_line.text()
        cargo = self.cargo_box.currentText()
        # faltando lógica para verificar se usuário já existe no banco com o mesmo nome.
        print(user,name,passw,email,cargo)
    
    def eventFilter(self, obj, event):

        if obj == self.btn_search:
            if event.type() == QtCore.QEvent.Enter:
                self.search_frame.setStyleSheet("QPushButton#btnSearch{border:2px solid #666666;border-radius:15px;}")
            elif event.type() == QtCore.QEvent.Leave:
                self.search_frame.setStyleSheet("QPushButton#btnSearch{border: 0px solid transparent;}")

        elif obj == self.btn_cad:
            if event.type() == QtCore.QEvent.Enter:
                self.cad_frame.setStyleSheet("QPushButton#btnCad{border:2px solid #666666;border-radius:15px;}")
            elif event.type() == QtCore.QEvent.Leave:
                self.cad_frame.setStyleSheet("QPushButton#btnCad{border: 0px solid transparent;}")

        elif obj == self.btn_info:
            if event.type() == QtCore.QEvent.Enter:
                self.info_frame.setStyleSheet("QPushButton#btnDetails{border:2px solid #666666;border-radius:15px;}")
            elif event.type() == QtCore.QEvent.Leave:
                self.info_frame.setStyleSheet("QPushButton#btnDetails{border: 0px solid transparent;}")

        return super().eventFilter(obj, event)
    

    def delete(self):
        print('Delete')
    
    def edit(self):
        print('Edit')
    
    def info(self):
        print('Info')
    
    def search(self):
        print('Search')
        
    def clear(self):
        self.user_edit_line = self.findChild(QLineEdit, "userEdit")
        self.name_edit_line = self.findChild(QLineEdit, "nameEdit")
        self.pass_edit_line = self.findChild(QLineEdit, "passEdit")
        self.email_edit_line = self.findChild(QLineEdit, "emailEdit")
        self.user_edit_line.clear() 
        self.name_edit_line.clear()
        self.pass_edit_line.clear()
        self.email_edit_line.clear()
        print('Teste')

class user_info(QWidget):
    def __init__(self):
        super().__init__()
        self.user_info = uic.loadUi("templates/interfaces/user_info.ui", self)
        self.user_btn_details = self.findChild(QPushButton, "btn_details")
        self.user_btn_details.clicked.connect(self.user_details)

    def user_details(self):
        self.user_d = user_details()
        self.user_frame = self.findChild(QFrame, "frame_user_details")
        self.f_info = self.findChild(QFrame, "frame_info")
        self.user_frame.layout().addWidget(self.user_d)
        self.f_info.hide()
        self.user_frame.show()

class bag_view(QWidget):
    def __init__(self):
        super().__init__()
        self.bag_screen = uic.loadUi("templates/interfaces/bag.ui", self)
        self.produtos = []
        self.produtos_temporarios = []


        self.btn_add_item = self.findChild(QPushButton, "regItem")
        self.btn_add_item.clicked.connect(self.bag_cad)
        self.btn_add_item.installEventFilter(self)


        self.btn_details = self.findChild(QPushButton, "details_btn")
        self.btn_details.installEventFilter(self)
        self.btn_details.hide()


        self.btn_del1 = self.findChild(QPushButton, "del_btn")
        self.btn_del1.installEventFilter(self)

        self.frame_view = self.findChild(QFrame, "frame_view")
        self.del_frame = self.findChild(QFrame, "del_frame")
        self.frame_search = self.findChild(QFrame, "frame_search")
        self.btn_top_frame = self.findChild(QFrame, "btn_top_frame")


        self.shadow_view = QGraphicsDropShadowEffect() 
        self.shadow_view.setOffset(0, 0)
        self.shadow_view.setBlurRadius(9)
        self.shadow_view.setColor(QtGui.QColor(0, 0, 0, 128))


        self.shadow_del = QGraphicsDropShadowEffect() 
        self.shadow_del.setOffset(0, 0)
        self.shadow_del.setBlurRadius(9)
        self.shadow_del.setColor(QtGui.QColor(0, 0, 0, 128))


        self.shadow_search = QGraphicsDropShadowEffect() 
        self.shadow_search.setOffset(0, 0)
        self.shadow_search.setBlurRadius(9)
        self.shadow_search.setColor(QtGui.QColor(0, 0, 0, 128))


        self.shadow_btn_top = QGraphicsDropShadowEffect() 
        self.shadow_btn_top.setOffset(0, 0)
        self.shadow_btn_top.setBlurRadius(9)
        self.shadow_btn_top.setColor(QtGui.QColor(0, 0, 0, 128))

        self.frame_view.setGraphicsEffect(self.shadow_view)
        self.frame_search.setGraphicsEffect(self.shadow_search)
        self.del_frame.setGraphicsEffect(self.shadow_del)
        self.btn_top_frame.setGraphicsEffect(self.shadow_btn_top)

        #efetua a pesquisa usando a storec proc st_pesquisar
        self.modelo = conecta_procedure_tela('st_pesquisar', '')
        self.table_item = self.findChild(QTableView, "table_bag")
        self.table_item.setModel(self.modelo)
        self.edtPesquisa = self.findChild(QLineEdit, "line_search")
        self.edtPesquisa.textChanged.connect(lambda text: self.pesquisa(text))
        
        self.table_item.verticalHeader().setVisible(False)
        self.table_item.resizeColumnsToContents()
        self.table_item.setColumnWidth(1, 250)
        
        header = self.table_item.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        
        self.table_item.setAlternatingRowColors(True)
        self.table_item.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        
        self.table_item.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_item.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.selected_rows = []
        self.table_item.clicked.connect(self.handle_row_click)
        
        self.l_t = None
        if self.l_t != None:
            self.btn_details.show()
        else:
            self.btn_details.hide()
#retirada redundancia que estava acontecendo nessa função pesquisa
    def pesquisa(self,pesquisado):
            self.modelo = conecta_procedure_tela('st_pesquisar', pesquisado)
            self.table_item.setModel(self.modelo)
   
    def eventFilter(self, obj, event):
        pass
        return super().eventFilter(obj, event)


    def lista_itens(self):
        pass


    def bag_cad(self):
        self.cad_itens = bag_item_cad()
        self.cad_frame = self.findChild(QFrame, "cad_frame")
        self.body_frame = self.findChild(QFrame, "body_frame")
        self.cad_frame.layout().addWidget(self.cad_itens)
        self.btn_teste = self.findChild(QPushButton, "test")
        self.body_frame.hide()
        self.cad_frame.show()        


    '''def del_item(self, item): #apagar?
        self.produtos_temporarios.remove(item)
        item.deleteLater()
        self.layout_tb = self.table_item.layout()
        self.layout_tb.removeWidget(item)
        self.layout_tb.update()'''


    def handle_row_click(self, index):
        row = index.row()
        print(f"Dados da linha {row}:")

        for column in range(self.table_item.model().columnCount()):
            data = self.table_item.model().index(row, column).data()
            id_patrimonio = self.table_item.model().index(row, 0).data()
            print(f"{self.table_item.model().horizontalHeaderItem(column).text()}: {data}")
            if self.table_item.model().horizontalHeaderItem(column).text() == 'ID':
                self.l_t = data
        self.selected_rows.clear()
        self.selected_rows.append(self.table_item.model().index(row, column).data())
        self.btn_details.show()
        self.btn_details.clicked.connect(self.details_screen)
        print('idptr', id_patrimonio)
        self.btn_del1.clicked.connect(lambda: deletar_patrimonio(id_patrimonio))
        

    def details_screen(self):
        self.window_details = detail_window(str(self.l_t), str(self.l_t))
        self.window_details.exec_()
        print('Faltando ajuste')

class bag_item_cad(QWidget):
    def __init__(self):
        super().__init__()
        
        self.cad_item = uic.loadUi("templates/interfaces/bag_item_cad.ui", self)
        self.name_item = self.findChild(QLineEdit, "name_prod");        
        self.quantidade = self.findChild(QSpinBox, "qnt")
        self.cbxCategoria = self.findChild(QComboBox, "cbxCategoria")
        self.dteDataRecebimento = self.findChild(QDateEdit, "date_rec")
        self.cbxSetoresResponsaveis = self.findChild(QComboBox, "cbxSetResp")
        self.cbxSituacao = self.findChild(QComboBox, "cbxSituacao")
        
        self.edtChaveAcesso = self.findChild(QLineEdit, "edtChaveAcessoNota")
        self.edtNumero = self.findChild(QLineEdit, "edtNumeroNota")
        self.edtSerie = self.findChild(QLineEdit, "edtSerieNota")
        self.cbxFornecedores = self.findChild(QComboBox, "cbxFornecedor")
        self.dteDataAquisicao = self.findChild(QDateEdit, "date_buy")
        
        #colocando as categorias na combo box
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute('select nome from categorias order by nome')
        resultado = cursor.fetchall()
        for dados in resultado:
            self.cbxCategoria.addItem(dados[0])
            
        cursor.execute('select nome from fornecedores order by nome')
        resultado_forn = cursor.fetchall()
        for dados in resultado_forn:
            self.cbxFornecedor.addItem(dados[0])
            
        cursor.execute('select nome from setores_responsaveis order by nome')
        resultado_set_resp = cursor.fetchall()
        for dados in resultado_set_resp:
            self.cbxSetoresResponsaveis.addItem(dados[0])
            
        cursor.execute('select nome from situacoes order by nome')
        resultado_situacoes = cursor.fetchall()
        for dados in resultado_situacoes:
            self.cbxSituacao.addItem(dados[0])
            
        cursor.close()
        fechar_conexao(con)

        # filtro para o campo de valor, aplicada regras para não aceitar virgula e nem outros digitos.
        self.number_input = self.findChild(QLineEdit, "value_prod")
        validator = QDoubleValidator(0.0, 10000.0, 2)
        validator.setNotation(QDoubleValidator.StandardNotation)
        validator.setLocale(self.number_input.locale())
        locale = QLocale(QLocale.C)
        validator.setLocale(locale)
        self.number_input.setValidator(validator)
        self.btn_ok = self.findChild(QPushButton, "ok_btn")
        self.number_input.keyPressEvent = self.keyPressEvent
        self.btn_confirm = self.findChild(QPushButton, "cadItens")
        self.btn_confirm.clicked.connect(self.confirm)

        self.btn_del2 = self.findChild(QPushButton, "del")
        self.btn_del2.clicked.connect(self.deletar_item)

        self.listagem = {}
        self.id_counter = 0
        self.lista_produtos = self.findChild(QTableView, "list_itens_add")

        self.btn_ok.clicked.connect(self.temp_list)

        self.selected_row = None

        self.lista_produtos.clicked.connect(self.handle_row_click)
        self.selected_rows = []
        self.lista_produtos.setGridStyle(Qt.NoPen)
        self.lista_produtos.setAlternatingRowColors(True)
        self.lista_produtos.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        self.lista_produtos.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.lista_produtos.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.value_result = self.findChild(QLabel, "value_r")

        self.number_input.textChanged.connect(self.update_label)
        self.quantidade.valueChanged.connect(self.update_label)

        self.atualizar_tabela()

    def update_label(self):
        try:
            valor = float(self.number_input.text() if self.number_input.text() else 0)
            q = float(self.quantidade.value())
            v = float(valor)
            resultado = q * v
            self.value_result.setText(f"Valor Total: {resultado:.2f}")
        except ValueError:
            self.value_result.setText("Por favor, insira um número válido no campo de valor.")
        pass
    

    def temp_list(self):        
        nome = self.name_item.text().strip()
        valor = self.number_input.text()
        categoria = self.cat_sel_nome = self.cbxCategoria.currentText()
        quantidade = self.quantidade.value()


        if not nome or not valor or quantidade == 0:
            print('Faltando valores. Verifique!')
        else:
            self.id_counter += 1
            self.listagem[self.id_counter] = [nome, valor, categoria, quantidade]
            self.atualizar_tabela()
            self.clear_c()
        
    def clear_c(self):
        self.name_item.clear()
        self.number_input.clear()
        self.quantidade.clear()
        self.quantidade.setValue(0)

    def atualizar_tabela(self):
        mdl = QStandardItemModel()
        hdr = ["Nome", "Valor", "Categoria", "Quantidade"]
        mdl.setHorizontalHeaderLabels(hdr)
        for item_id, item_data in self.listagem.items():
            row_items = []
            for value in item_data:
                item = QStandardItem(str(value))
                row_items.append(item)
            mdl.appendRow(row_items)
        self.lista_produtos.setModel(mdl)
        self.lista_produtos.verticalHeader().setVisible(False)
        self.lista_produtos.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.lista_produtos.horizontalHeader().setStretchLastSection(True)
        print(self.listagem)
        
                
    # função responsável pela interação de seleção de linha
    # print com o for apenas para teste
    def handle_row_click(self, index):
        row = index.row()
        self.selected_row = row
        print(f"Dados da linha {row}:")
        for column in range(self.lista_produtos.model().columnCount()):
            data = self.lista_produtos.model().index(row, column).data()
            print(f"{self.lista_produtos.model().horizontalHeaderItem(column).text()}: {data}")
        self.selected_rows.clear()
        self.selected_rows.append(self.lista_produtos.model().index(row, column).data())
        
    def deletar_item(self):
        if self.selected_row is None:
            print("Nenhuma linha selecionada para deletar.")
            return
        item_id = list(self.listagem.keys())[self.selected_row]
        del self.listagem[item_id]
        self.atualizar_tabela()
        self.selected_row = None
        if self.listagem:
            self.id_counter = max(self.listagem.keys())
        else:
            self.id_counter = 0

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == Qt.Key_Comma:
            return
        QLineEdit.keyPressEvent(self.number_input, event)
        
    def confirm(self):
        print("Confirmando os itens:", self.listagem)
        for item_id, item_data in self.listagem.items():
            print(f'Produto id:{item_id}, Produto da lista: {item_data}')
            
        data_aquisicao = self.dteDataAquisicao.date().toString("yyyy-MM-dd")
        chave_acesso = self.edtChaveAcesso.text()
        numero = self.edtNumero.text()
        serie = self.edtSerie.text()
        
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute('select idcategoria from categorias where nome = %s', (self.cbxCategoria.currentText(),))
        resultado_cat = cursor.fetchone()
        self.cat_sel_id = resultado_cat[0]
        
        cursor.execute('select idsetor from setores_responsaveis where nome = %s', (self.cbxSetoresResponsaveis.currentText(),))
        resultado_set_resp = cursor.fetchone()
        self.set_resp_sel_id = resultado_set_resp[0]
        
        cursor.execute('select idsituacao from situacoes where nome = %s', (self.cbxSituacao.currentText(),))
        resultado_situacao = cursor.fetchone()
        self.sit_sel_id = resultado_situacao[0]
        
        cursor.execute('select idfornecedor from fornecedores where nome = %s', (self.cbxFornecedores.currentText(),))
        resultado_forn = cursor.fetchone()
        self.forn_sel_id = resultado_forn[0]
        

        nota_sel_id = cadastra_nota(chave_acesso, numero, serie, self.forn_sel_id, data_aquisicao, )

        data_recebimento = self.dteDataRecebimento.date().toString("yyyy-MM-dd")
        
        
        try:
            for item_id, item_data in self.listagem.items():
                nome, valor_unitario, categoria, quantidade = item_data
            
            cursor.execute('set @idusuario = %s', (id_user,))
            
            cursor.callproc('cadastra_quantidade', [nome, valor_unitario, data_recebimento, nota_sel_id, self.cat_sel_id, self.set_resp_sel_id, 1, self.sit_sel_id, quantidade])
            con.commit()
            print('Produto(s) cadastrado(s) com sucesso!', nome, valor_unitario, data_recebimento, nota_sel_id, self.cat_sel_id, self.set_resp_sel_id, self.sit_sel_id, quantidade)
        except Exception as e:
            print('Erro no cadastro', e)
        finally:
            cursor.close()   
            fechar_conexao(con)
    
        self.listagem.clear()
        self.atualizar_tabela()
        print("Itens confirmados e adicionados ao banco de dados:", self.listagem)

class user_details(QWidget):
    def __init__(self):
        super().__init__()
        self.user_details = uic.loadUi("templates/interfaces/user_details.ui", self)
        
class user_details(QWidget):
    def __init__(self):
        super().__init__()
        self.user_details = uic.loadUi("templates/interfaces/user_details.ui", self)
        
class items_view(QWidget):
    def __init__(self, interface):
        super().__init__()
        self.interface = interface
        self.item_view = uic.loadUi("templates/interfaces/item_view.ui", self)
        self.btn_categ = self.findChild(QPushButton, "cat_item_btn")
        self.btn_categ.clicked.connect(self.categ_view)

    def categ_view(self):
        self.categview = categ_view()
        self.categview.configRequested.connect(self.interface.config_screen)
        self.frame_v1 = self.findChild(QFrame, "body_item")
        self.frame_v2 = self.findChild(QFrame, "frame_view2")
        self.frame_v2.layout().addWidget(self.categview)
        self.frame_v2.show()
        self.frame_v1.hide()
        


class categ_view(QWidget):   
    configRequested = pyqtSignal()
    def __init__(self):
        super().__init__()
        self.item_category = uic.loadUi("templates/interfaces/categ_view.ui", self)
        self.btn_config = self.findChild(QPushButton, "category")
        self.btn_config.clicked.connect(self.on_config_click)
    def on_config_click(self):
        self.configRequested.emit()




class rel_view(QWidget):
    def __init__(self):
        super().__init__()
        self.rel_view = uic.loadUi("templates/interfaces/rel_view.ui", self)


class patr_view(QWidget):
    def __init__(self):
        super().__init__()
        self.patr_view = uic.loadUi("templates/interfaces/patr_view.ui", self)
        self.list_ptr = self.findChild(QTableView, "list_patr")
        self.grap_view = self.findChild(QGraphicsView, "grap_view")
        self.date_i = self.findChild(QDateEdit, "dt_inicial")
        self.date_f = self.findChild(QDateEdit, "dt_final")
        self.btn_filter = self.findChild(QPushButton, "filter_dt")
        self.btn_filter.clicked.connect(self.filter_data)
        self.btn_rlt = self.findChild(QPushButton, "rlt_btn")
        self.btn_rlt.clicked.connect(self.rlt_patrimonio)
        self.set_default_dates()
        self.setup_table()
        self.setup_graph()
        self.create_graph(self.date_i.date().toString("yyyy-MM-dd"), self.date_f.date().toString("yyyy-MM-dd"))
        self.filter_data()

    def set_default_dates(self):
        today = datetime.now()
        three_years_ago = today - timedelta(days=365 * 3)
        self.date_i.setDate(three_years_ago)
        self.date_f.setDate(today)

    def setup_table(self):
        self.list_ptr.setEditTriggers(QTableView.NoEditTriggers)
        self.list_ptr.setSelectionMode(QTableView.NoSelection)
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(["Nome", "Valor Uni.", "Data de Recebimento"])
        self.list_ptr.setModel(self.model)
        header = self.list_ptr.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        self.list_ptr.verticalHeader().setVisible(False)
        self.list_ptr.setGridStyle(Qt.NoPen)
        self.list_ptr.setAlternatingRowColors(True)
        self.list_ptr.setStyleSheet("alternate-background-color: #F0F0F0; background-color: #FFFFFF;")
        self.load_table_data()

    def load_table_data(self, d_i=None, d_f=None):
        self.model.clear()
        con = criar_conexao()
        cursor = con.cursor()

        if d_i and d_f:
            cursor.execute("""
                SELECT nome, valor_unitario, data_recebimento 
                FROM patrimonios 
                WHERE data_recebimento BETWEEN %s AND %s
            """, (d_i, d_f))
        else:
            cursor.execute("SELECT nome, valor_unitario, data_recebimento FROM patrimonios")            

        for row in cursor.fetchall():
            items = [QStandardItem(str(cell)) for cell in row]
            for item in items:
                item.setSelectable(False)
            self.model.appendRow(items)
        self.model.setHorizontalHeaderLabels(["Nome", "Valor Uni.", "Data de Recebimento"])
        self.list_ptr.setModel(self.model)
        self.list_ptr.resizeColumnsToContents()
        self.list_ptr.horizontalHeader().setStretchLastSection(True)

    def setup_graph(self):
        layout = QVBoxLayout()
        self.figure = plt.figure(figsize=(8, 6))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)
        self.grap_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        widget = QWidget()
        widget.setLayout(layout)
        self.scene = QGraphicsScene()
        self.scene.addWidget(widget)
        self.grap_view.setScene(self.scene)

    def filter_data(self):
        d_i = self.date_i.date().toString("yyyy-MM-dd")
        d_f = self.date_f.date().toString("yyyy-MM-dd")
        self.load_table_data(d_i, d_f)
        self.create_graph(d_i, d_f)

    def create_graph(self, d_i, d_f):
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute("""
            SELECT DATE(data_recebimento) as data, 
                SUM(valor_unitario) as valor_total
            FROM patrimonios 
            WHERE data_recebimento BETWEEN %s AND %s
            GROUP BY DATE(data_recebimento)
            ORDER BY data_recebimento
        """, (d_i, d_f))

        data = cursor.fetchall()
        df = pd.DataFrame(data, columns=['data', 'valor_total'])

        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.plot(df['data'], df['valor_total'], color='green', marker='o')
        ax.set_title('Valor Total por Data')
        ax.set_xlabel('Data')
        ax.set_ylabel('Valor Total (R$)')
        ax.tick_params(axis='x', rotation=45)
        self.figure.tight_layout()
        self.canvas.draw()
        self.grap_view.fitInView(self.grap_view.sceneRect(), Qt.KeepAspectRatio)

    def rlt_patrimonio(self):
        d_i = self.date_i.date().toString("yyyy-MM-dd")
        d_f = self.date_f.date().toString("yyyy-MM-dd")

        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute("""
            SELECT nome, valor_unitario, num_patrimonio, data_recebimento 
            FROM patrimonios
            WHERE data_recebimento BETWEEN %s AND %s
        """, (d_i, d_f))
        data = cursor.fetchall()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Relatório de Patrimônio"

        bold_font = Font(bold=True)
        centered_alignment = Alignment(horizontal='center', vertical='center')

        worksheet['A1'] = 'Nome'
        worksheet['B1'] = 'Valor Unitário'
        worksheet['C1'] = 'Número de Patrimônio'
        worksheet['D1'] = 'Data de Recebimento'
        for col in range(1, 5):
            worksheet.cell(row=1, column=col).font = bold_font
            worksheet.cell(row=1, column=col).alignment = centered_alignment

        row = 2
        for nome, valor_unitario, num_patrimonio, data_recebimento in data:
            worksheet.cell(row=row, column=1, value=nome)
            worksheet.cell(row=row, column=2, value=valor_unitario)
            worksheet.cell(row=row, column=3, value=num_patrimonio)
            worksheet.cell(row=row, column=4, value=data_recebimento)
            row += 1

        # Salvar o arquivo
        file_path, _ = QFileDialog.getSaveFileName(self, "Salvar Relatório", os.path.expanduser("~"), "Arquivos Excel (*.xlsx)")
        if file_path:
            workbook.save(file_path)
            QMessageBox.information(self, "Relatório Gerado", f"O relatório foi salvo em: {file_path}")

class logs_view(QWidget):
    def __init__(self):
        super().__init__()
        self.logs_view = uic.loadUi("templates/interfaces/logs_view.ui", self)

class config_view(QWidget):
    def __init__(self):
        super().__init__()
        self.config_view = uic.loadUi("templates/interfaces/config_screen.ui", self)

class local_info(QWidget):
    def __init__(self):
        super().__init__()
        self.local_info = uic.loadUi("templates/interfaces/local_info.ui", self)
        
class detail_window(QDialog):
    def __init__(self, msg, type_msg):
        super().__init__()
        self.type_msg = type_msg
        self.msg = msg
        print(type_msg)
        self.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint)
        self.setModal(True)
        self.setStyleSheet("background-color: #f0f0f0; border: 1px solid #ccc;")
        layout = QVBoxLayout()
        con = criar_conexao()
        cursor = con.cursor()
        cursor.execute("SELECT * FROM patrimonios WHERE idpatrimonio = %s", (self.type_msg,))
        data = cursor.fetchone()
        label_mensagem = QLabel("Dados selecionados:" + str(data))
        label_mensagem.setAlignment(Qt.AlignCenter)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(10, 100, 10, 100)
        layout.setSpacing(10)
        layout.addWidget(label_mensagem)
        self.setLayout(layout)
        

class config_cargo(QWidget):
    def __init__(self):
        super().__init__()
        self.config_cargo = uic.loadUi("templates/interfaces/cargo_config.ui", self)
        
class config_cat(QWidget):
    def __init__(self):
        super().__init__()
        self.config_cat = uic.loadUi("templates/interfaces/categ_config.ui", self)
        
class config_forn(QWidget):
    def __init__(self):
        super().__init__()
        self.config_forn = uic.loadUi("templates/interfaces/forn_config.ui", self)
        
class config_local(QWidget):
    def __init__(self):
        super().__init__()
        self.config_local = uic.loadUi("templates/interfaces/local_config.ui", self)